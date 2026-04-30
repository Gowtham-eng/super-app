from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Query, UploadFile, File, BackgroundTasks
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict, EmailStr
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import ipaddress

from services.email_service import send_email, build_access_request_email, build_request_status_email, build_sync_report_email
from services.adrenalin_sync import sync_employees
from services.kissflow_scim_client import sync_to_kissflow, push_single_user_to_kissflow, get_kissflow_scim_config, save_kissflow_scim_config, resolve_managers_in_kissflow
from routes import scim as scim_router_module

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Configuration
JWT_SECRET = os.environ.get('JWT_SECRET', 'kissflow-iam-secret-key-2024')
JWT_ALGORITHM = 'HS256'
JWT_EXPIRATION_HOURS = 720  # 30 days

# Public URL - MUST be set for SAML SSO to work in Kubernetes
PUBLIC_URL = os.environ.get('PUBLIC_URL', '').rstrip('/')

app = FastAPI(title="Kissflow IAM - Identity & Access Management")
api_router = APIRouter(prefix="/api")
security = HTTPBearer()

# APScheduler for midnight HR sync
from apscheduler.schedulers.asyncio import AsyncIOScheduler
scheduler = AsyncIOScheduler()

async def scheduled_hr_sync():
    """Run Adrenalin HR sync for all organizations that have it configured"""
    logger = logging.getLogger("hr_sync")
    logger.info("Starting scheduled HR sync...")
    orgs = await db.organizations.find({"adrenalin_sync_enabled": True}, {"_id": 0}).to_list(100)
    if not orgs:
        # Default: sync for Refex org
        orgs = await db.organizations.find({}, {"_id": 0}).to_list(1)
    for org in orgs:
        try:
            result = await sync_employees(db, org["id"])
            # Log the HR sync
            await db.hr_sync_logs.insert_one({
                "org_id": org["id"],
                "timestamp": datetime.now(timezone.utc).isoformat(),
                "result": result,
            })
            # Log the Kissflow push result separately if present
            kf_result = result.get("kissflow_sync")
            if kf_result and not kf_result.get("skipped"):
                await db.kissflow_sync_logs.insert_one({
                    "org_id": org["id"],
                    "trigger_type": "scheduled_hr_sync",
                    "timestamp": datetime.now(timezone.utc).isoformat(),
                    "result": kf_result,
                })
            # Email report to admins
            admins = await db.users.find(
                {"org_id": org["id"], "role": "org_admin", "status": {"$ne": "disabled"}},
                {"_id": 0, "email": 1}
            ).to_list(100)
            admin_emails = [a["email"] for a in admins if a.get("email")]
            if admin_emails and (result["created"] > 0 or result["disabled"] > 0 or result["errors"]):
                html = build_sync_report_email(result["created"], result["disabled"], result["total"], result["errors"])
                await send_email(admin_emails, "HR Sync Report - Refex Super App", html)
            logger.info(f"Sync complete for org {org['id']}: {result}")
        except Exception as e:
            logger.error(f"Sync failed for org {org['id']}: {e}")

@app.on_event("startup")
async def start_scheduler():
    # Fix _id issue for hr_sync_logs
    scheduler.add_job(scheduled_hr_sync, 'cron', hour=0, minute=0, id='hr_sync_midnight')
    scheduler.start()
    logging.getLogger("hr_sync").info("Scheduler started - HR sync at midnight daily")

# Static uploads directory
UPLOAD_DIR = ROOT_DIR / 'uploads'
UPLOAD_DIR.mkdir(exist_ok=True)
from fastapi.staticfiles import StaticFiles
app.mount("/api/uploads", StaticFiles(directory=str(UPLOAD_DIR)), name="uploads")

# ===================== MODELS =====================

# Organization Models
class OrganizationCreate(BaseModel):
    name: str
    domain: str
    description: Optional[str] = None

class Organization(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    domain: str
    description: Optional[str] = None
    created_at: str
    status: str = "active"

# Permission & Role Models
class Permission(BaseModel):
    id: str
    name: str
    description: str
    resource: str  # apps, users, groups, settings, audit
    actions: List[str]  # create, read, update, delete, manage

class RoleCreate(BaseModel):
    name: str
    description: Optional[str] = None
    permissions: List[str]  # permission IDs
    org_id: str

class Role(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: Optional[str] = None
    permissions: List[str]
    org_id: str
    is_system: bool = False
    created_at: str

# Group Models
class GroupCreate(BaseModel):
    name: str
    description: Optional[str] = None
    org_id: str
    parent_id: Optional[str] = None
    role_ids: List[str] = []

class Group(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: Optional[str] = None
    org_id: str
    parent_id: Optional[str] = None
    role_ids: List[str] = []
    member_count: int = 0
    created_at: str

# User Models
class UserCreate(BaseModel):
    email: EmailStr
    password: str
    name: str
    org_id: str

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    status: Optional[str] = None
    group_ids: Optional[List[str]] = None
    role_ids: Optional[List[str]] = None

# Application Models (SAML & OIDC)
class SAMLAppCreate(BaseModel):
    name: str
    description: Optional[str] = None
    org_id: str
    entity_id: str
    acs_url: str
    slo_url: Optional[str] = None
    home_url: Optional[str] = None
    name_id_format: str = "urn:oasis:names:tc:SAML:1.1:nameid-format:emailAddress"
    sign_assertions: bool = True
    sign_response: bool = True
    attribute_mappings: Dict[str, str] = {}
    logo_url: Optional[str] = None
    allowed_group_ids: List[str] = []
    allowed_role_ids: List[str] = []

class OIDCAppCreate(BaseModel):
    name: str
    description: Optional[str] = None
    org_id: str
    redirect_uris: List[str]
    logout_uris: List[str] = []
    scopes: List[str] = ["openid", "profile", "email"]
    grant_types: List[str] = ["authorization_code"]
    logo_url: Optional[str] = None
    home_url: Optional[str] = None
    allowed_group_ids: List[str] = []
    allowed_role_ids: List[str] = []

# Access Policy Models
class AccessPolicyCreate(BaseModel):
    name: str
    description: Optional[str] = None
    org_id: str
    app_ids: List[str] = []  # Empty means all apps
    conditions: Dict[str, Any] = {}  # ip_whitelist, ip_blacklist, time_restrictions, mfa_required

class AccessPolicy(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    name: str
    description: Optional[str] = None
    org_id: str
    app_ids: List[str] = []
    conditions: Dict[str, Any] = {}
    enabled: bool = True
    created_at: str

# Access Request Models
class AccessRequestCreate(BaseModel):
    app_id: str
    reason: str

class AccessRequest(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    user_id: str
    user_email: str
    user_name: str
    app_id: str
    app_name: str
    app_type: str  # saml or oidc
    reason: str
    status: str = "pending"  # pending, approved, rejected
    reviewed_by: Optional[str] = None
    reviewed_at: Optional[str] = None
    org_id: str
    created_at: str

# Audit Log Models
class AuditLog(BaseModel):
    model_config = ConfigDict(extra="ignore")
    id: str
    org_id: str
    user_id: Optional[str] = None
    user_email: Optional[str] = None
    action: str  # login, logout, app_access, user_created, role_changed, etc.
    resource_type: str  # user, app, group, role, policy
    resource_id: Optional[str] = None
    details: Dict[str, Any] = {}
    ip_address: Optional[str] = None
    user_agent: Optional[str] = None
    timestamp: str
    status: str = "success"  # success, failure

# ===================== HELPERS =====================

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())

def create_token(user_id: str, email: str, org_id: str, role: str) -> str:
    payload = {
        'user_id': user_id,
        'email': email,
        'org_id': org_id,
        'role': role,
        'exp': datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS),
        'iat': datetime.now(timezone.utc)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)

def decode_token(token: str) -> dict:
    try:
        return jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")

async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    token = credentials.credentials
    payload = decode_token(token)
    user = await db.users.find_one({"id": payload['user_id']}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    return user

async def log_audit(org_id: str, action: str, resource_type: str, user_id: str = None, 
                   user_email: str = None, resource_id: str = None, details: dict = None,
                   ip_address: str = None, status: str = "success"):
    audit_doc = {
        "id": str(uuid.uuid4()),
        "org_id": org_id,
        "user_id": user_id,
        "user_email": user_email,
        "action": action,
        "resource_type": resource_type,
        "resource_id": resource_id,
        "details": details or {},
        "ip_address": ip_address,
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": status
    }
    await db.audit_logs.insert_one(audit_doc)

def get_public_base_url(request: Request = None) -> str:
    """Get the public-facing base URL. Uses PUBLIC_URL env var, falls back to request headers."""
    if PUBLIC_URL:
        return PUBLIC_URL
    if request:
        forwarded_host = request.headers.get('x-forwarded-host')
        host = forwarded_host or request.headers.get('host', '')
        scheme = request.headers.get('x-forwarded-proto', 'https')
        return f"{scheme}://{host}"
    return 'http://localhost:8001'

def generate_self_signed_cert():
    from cryptography import x509
    from cryptography.x509.oid import NameOID
    from cryptography.hazmat.primitives import hashes, serialization
    from cryptography.hazmat.primitives.asymmetric import rsa
    
    key = rsa.generate_private_key(public_exponent=65537, key_size=2048)
    subject = issuer = x509.Name([
        x509.NameAttribute(NameOID.COUNTRY_NAME, "US"),
        x509.NameAttribute(NameOID.ORGANIZATION_NAME, "Kissflow IAM"),
        x509.NameAttribute(NameOID.COMMON_NAME, "iam.kissflow.local"),
    ])
    cert = (
        x509.CertificateBuilder()
        .subject_name(subject)
        .issuer_name(issuer)
        .public_key(key.public_key())
        .serial_number(x509.random_serial_number())
        .not_valid_before(datetime.now(timezone.utc))
        .not_valid_after(datetime.now(timezone.utc) + timedelta(days=365))
        .sign(key, hashes.SHA256())
    )
    cert_pem = cert.public_bytes(serialization.Encoding.PEM).decode()
    key_pem = key.private_bytes(
        serialization.Encoding.PEM,
        serialization.PrivateFormat.TraditionalOpenSSL,
        serialization.NoEncryption()
    ).decode()
    return cert_pem, key_pem

def generate_saml_metadata(app: dict, base_url: str) -> str:
    cert = app.get('certificate', '')
    cert_clean = cert.replace('-----BEGIN CERTIFICATE-----', '').replace('-----END CERTIFICATE-----', '').replace('\n', '').strip()
    
    return f'''<?xml version="1.0" encoding="UTF-8"?>
<md:EntityDescriptor xmlns:md="urn:oasis:names:tc:SAML:2.0:metadata" 
                     xmlns:ds="http://www.w3.org/2000/09/xmldsig#"
                     entityID="{app['entity_id']}">
    <md:IDPSSODescriptor WantAuthnRequestsSigned="true" 
                         protocolSupportEnumeration="urn:oasis:names:tc:SAML:2.0:protocol">
        <md:KeyDescriptor use="signing">
            <ds:KeyInfo><ds:X509Data><ds:X509Certificate>{cert_clean}</ds:X509Certificate></ds:X509Data></ds:KeyInfo>
        </md:KeyDescriptor>
        <md:NameIDFormat>{app['name_id_format']}</md:NameIDFormat>
        <md:SingleSignOnService Binding="urn:oasis:names:tc:SAML:2.0:bindings:HTTP-POST" Location="{base_url}/api/saml/{app['id']}/sso"/>
        <md:SingleLogoutService Binding="urn:oasis:names:tc:SAML:2.0:bindings:HTTP-Redirect" Location="{base_url}/api/saml/{app['id']}/slo"/>
    </md:IDPSSODescriptor>
    <md:Organization>
        <md:OrganizationName xml:lang="en">{app['name']}</md:OrganizationName>
        <md:OrganizationDisplayName xml:lang="en">{app['name']}</md:OrganizationDisplayName>
        <md:OrganizationURL xml:lang="en">{base_url}</md:OrganizationURL>
    </md:Organization>
</md:EntityDescriptor>'''

async def check_user_app_access(user: dict, app: dict) -> bool:
    """Check if user has access to an application based on direct approval, groups, or roles.
    Access must be explicitly granted - no implicit access for unrestricted apps."""
    user_id = user.get('id')
    user_groups = set(user.get('group_ids', []))
    user_roles = set(user.get('role_ids', []))
    
    allowed_groups = set(app.get('allowed_group_ids', []))
    allowed_roles = set(app.get('allowed_role_ids', []))
    approved_users = set(app.get('approved_user_ids', []))
    
    # Org admins always have access
    if user.get('role') == 'org_admin':
        return True
    
    # Check if user was directly approved for this app
    if user_id in approved_users:
        return True
    
    # Check group membership
    if allowed_groups and user_groups.intersection(allowed_groups):
        return True
    
    # Check role assignment
    if allowed_roles and user_roles.intersection(allowed_roles):
        return True
    
    # Check if user is in a group that has an allowed role
    for group_id in user_groups:
        group = await db.groups.find_one({"id": group_id}, {"_id": 0})
        if group:
            group_roles = set(group.get('role_ids', []))
            if group_roles.intersection(allowed_roles):
                return True
    
    return False

async def check_access_policies(user: dict, app: dict, request: Request) -> tuple:
    """Check access policies and return (allowed, reason)"""
    org_id = user.get('org_id')
    app_id = app.get('id')
    client_ip = request.client.host if request.client else None
    
    policies = await db.access_policies.find({
        "org_id": org_id,
        "enabled": True,
        "$or": [
            {"app_ids": []},
            {"app_ids": app_id}
        ]
    }, {"_id": 0}).to_list(100)
    
    for policy in policies:
        conditions = policy.get('conditions', {})
        
        # IP Whitelist check
        if 'ip_whitelist' in conditions and conditions['ip_whitelist']:
            if client_ip:
                allowed = False
                for ip_range in conditions['ip_whitelist']:
                    try:
                        if '/' in ip_range:
                            if ipaddress.ip_address(client_ip) in ipaddress.ip_network(ip_range, strict=False):
                                allowed = True
                                break
                        elif client_ip == ip_range:
                            allowed = True
                            break
                    except:
                        pass
                if not allowed:
                    return False, f"IP {client_ip} not in whitelist"
        
        # IP Blacklist check
        if 'ip_blacklist' in conditions and conditions['ip_blacklist']:
            if client_ip:
                for ip_range in conditions['ip_blacklist']:
                    try:
                        if '/' in ip_range:
                            if ipaddress.ip_address(client_ip) in ipaddress.ip_network(ip_range, strict=False):
                                return False, f"IP {client_ip} is blacklisted"
                        elif client_ip == ip_range:
                            return False, f"IP {client_ip} is blacklisted"
                    except:
                        pass
        
        # Time restrictions check
        if 'time_restrictions' in conditions:
            tr = conditions['time_restrictions']
            now = datetime.now(timezone.utc)
            
            if 'allowed_days' in tr:
                if now.strftime('%A').lower() not in [d.lower() for d in tr['allowed_days']]:
                    return False, "Access not allowed on this day"
            
            if 'start_hour' in tr and 'end_hour' in tr:
                current_hour = now.hour
                if not (tr['start_hour'] <= current_hour < tr['end_hour']):
                    return False, f"Access only allowed between {tr['start_hour']}:00 and {tr['end_hour']}:00 UTC"
    
    return True, None

# ===================== DEFAULT DATA SEEDING =====================

async def seed_default_permissions():
    """Seed default system permissions"""
    default_permissions = [
        {"id": "perm_apps_manage", "name": "Manage Applications", "description": "Create, edit, delete applications", "resource": "apps", "actions": ["create", "read", "update", "delete"]},
        {"id": "perm_apps_read", "name": "View Applications", "description": "View application list and details", "resource": "apps", "actions": ["read"]},
        {"id": "perm_users_manage", "name": "Manage Users", "description": "Create, edit, delete users", "resource": "users", "actions": ["create", "read", "update", "delete"]},
        {"id": "perm_users_read", "name": "View Users", "description": "View user list and details", "resource": "users", "actions": ["read"]},
        {"id": "perm_groups_manage", "name": "Manage Groups", "description": "Create, edit, delete groups", "resource": "groups", "actions": ["create", "read", "update", "delete"]},
        {"id": "perm_roles_manage", "name": "Manage Roles", "description": "Create, edit, delete roles", "resource": "roles", "actions": ["create", "read", "update", "delete"]},
        {"id": "perm_policies_manage", "name": "Manage Policies", "description": "Create, edit, delete access policies", "resource": "policies", "actions": ["create", "read", "update", "delete"]},
        {"id": "perm_audit_read", "name": "View Audit Logs", "description": "View audit logs and reports", "resource": "audit", "actions": ["read"]},
        {"id": "perm_requests_manage", "name": "Manage Access Requests", "description": "Approve or reject access requests", "resource": "requests", "actions": ["read", "update"]},
        {"id": "perm_org_manage", "name": "Manage Organization", "description": "Edit organization settings", "resource": "organization", "actions": ["read", "update"]},
    ]
    
    for perm in default_permissions:
        existing = await db.permissions.find_one({"id": perm['id']})
        if not existing:
            await db.permissions.insert_one(perm)

# ===================== AUTH ROUTES =====================

@api_router.post("/auth/register")
async def register(user: UserCreate, request: Request):
    existing = await db.users.find_one({"email": user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")
    
    # Verify organization exists
    org = await db.organizations.find_one({"id": user.org_id}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=400, detail="Organization not found")
    
    user_id = str(uuid.uuid4())
    user_count = await db.users.count_documents({"org_id": user.org_id})
    role = "org_admin" if user_count == 0 else "user"
    
    user_doc = {
        "id": user_id,
        "email": user.email,
        "password": hash_password(user.password),
        "name": user.name,
        "org_id": user.org_id,
        "role": role,
        "group_ids": [],
        "role_ids": [],
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.users.insert_one(user_doc)
    await log_audit(user.org_id, "user_registered", "user", user_id, user.email, user_id, 
                   {"name": user.name}, request.client.host if request.client else None)
    
    token = create_token(user_id, user.email, user.org_id, role)
    return {"token": token, "user": {"id": user_id, "email": user.email, "name": user.name, "role": role, "org_id": user.org_id}}

@api_router.post("/auth/login")
async def login(credentials: UserLogin, request: Request):
    user = await db.users.find_one({"email": credentials.email}, {"_id": 0})
    if not user or not verify_password(credentials.password, user['password']):
        raise HTTPException(status_code=401, detail="Invalid credentials")
    
    if user.get('status') != 'active':
        raise HTTPException(status_code=403, detail="Account is not active")
    
    await log_audit(user['org_id'], "user_login", "user", user['id'], user['email'], user['id'],
                   {}, request.client.host if request.client else None)
    
    token = create_token(user['id'], user['email'], user['org_id'], user['role'])
    return {"token": token, "user": {"id": user['id'], "email": user['email'], "name": user['name'], 
                                      "role": user['role'], "org_id": user['org_id']}}

@api_router.get("/auth/me")
async def get_me(user: dict = Depends(get_current_user)):
    org = await db.organizations.find_one({"id": user['org_id']}, {"_id": 0})
    return {**user, "organization": org}


# ===================== FILE UPLOAD =====================

@api_router.post("/upload/logo")
async def upload_logo(file: UploadFile = File(...), user: dict = Depends(get_current_user)):
    """Upload a logo image and return URL"""
    if not file.content_type or not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Only image files are allowed")
    
    ext = file.filename.rsplit('.', 1)[-1] if '.' in file.filename else 'png'
    filename = f"{uuid.uuid4().hex}.{ext}"
    filepath = UPLOAD_DIR / filename
    
    contents = await file.read()
    if len(contents) > 5 * 1024 * 1024:  # 5MB limit
        raise HTTPException(status_code=400, detail="File too large (max 5MB)")
    
    with open(filepath, 'wb') as f:
        f.write(contents)
    
    base_url = get_public_base_url()
    logo_url = f"{base_url}/api/uploads/{filename}"
    return {"logo_url": logo_url, "filename": filename}


@api_router.put("/users/me/profile-pic")
async def update_profile_pic(body: dict, user: dict = Depends(get_current_user)):
    """Update current user's profile picture URL"""
    profile_pic = body.get("profile_pic", "")
    await db.users.update_one({"id": user["id"]}, {"$set": {"profile_pic": profile_pic}})
    updated = await db.users.find_one({"id": user["id"]}, {"_id": 0, "password": 0})
    org = await db.organizations.find_one({"id": updated["org_id"]}, {"_id": 0})
    return {**updated, "organization": org}


# ===================== ORGANIZATION ROUTES =====================

@api_router.post("/organizations")
async def create_organization(org: OrganizationCreate):
    existing = await db.organizations.find_one({"domain": org.domain})
    if existing:
        raise HTTPException(status_code=400, detail="Domain already registered")
    
    org_id = str(uuid.uuid4())
    org_doc = {
        "id": org_id,
        "name": org.name,
        "domain": org.domain,
        "description": org.description,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.organizations.insert_one(org_doc)
    
    # Create default roles for the organization
    default_roles = [
        {"id": str(uuid.uuid4()), "name": "Administrator", "description": "Full access to all features", 
         "permissions": ["perm_apps_manage", "perm_users_manage", "perm_groups_manage", "perm_roles_manage", 
                        "perm_policies_manage", "perm_audit_read", "perm_requests_manage", "perm_org_manage"],
         "org_id": org_id, "is_system": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "name": "User Manager", "description": "Manage users and groups",
         "permissions": ["perm_users_manage", "perm_groups_manage", "perm_apps_read"],
         "org_id": org_id, "is_system": True, "created_at": datetime.now(timezone.utc).isoformat()},
        {"id": str(uuid.uuid4()), "name": "Viewer", "description": "Read-only access",
         "permissions": ["perm_apps_read", "perm_users_read"],
         "org_id": org_id, "is_system": True, "created_at": datetime.now(timezone.utc).isoformat()},
    ]
    for role in default_roles:
        await db.roles.insert_one(role)
    
    await seed_default_permissions()
    return {**org_doc, "_id": None}

@api_router.get("/organizations")
async def list_organizations():
    orgs = await db.organizations.find({}, {"_id": 0}).to_list(100)
    return orgs

@api_router.get("/organizations/{org_id}")
async def get_organization(org_id: str, user: dict = Depends(get_current_user)):
    if user['org_id'] != org_id and user['role'] != 'super_admin':
        raise HTTPException(status_code=403, detail="Access denied")
    org = await db.organizations.find_one({"id": org_id}, {"_id": 0})
    if not org:
        raise HTTPException(status_code=404, detail="Organization not found")
    return org

# ===================== ROLE ROUTES =====================

@api_router.get("/roles")
async def list_roles(user: dict = Depends(get_current_user)):
    roles = await db.roles.find({"org_id": user['org_id']}, {"_id": 0}).to_list(100)
    return roles

@api_router.post("/roles")
async def create_role(role: RoleCreate, request: Request, user: dict = Depends(get_current_user)):
    if role.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    role_id = str(uuid.uuid4())
    role_doc = {
        "id": role_id,
        "name": role.name,
        "description": role.description,
        "permissions": role.permissions,
        "org_id": role.org_id,
        "is_system": False,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.roles.insert_one(role_doc)
    await log_audit(user['org_id'], "role_created", "role", user['id'], user['email'], role_id,
                   {"name": role.name}, request.client.host if request.client else None)
    return {**role_doc, "_id": None}

@api_router.put("/roles/{role_id}")
async def update_role(role_id: str, update: dict, request: Request, user: dict = Depends(get_current_user)):
    role = await db.roles.find_one({"id": role_id, "org_id": user['org_id']}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if role.get('is_system'):
        raise HTTPException(status_code=400, detail="Cannot modify system roles")
    
    update.pop('id', None)
    update.pop('org_id', None)
    update.pop('is_system', None)
    await db.roles.update_one({"id": role_id}, {"$set": update})
    await log_audit(user['org_id'], "role_updated", "role", user['id'], user['email'], role_id,
                   update, request.client.host if request.client else None)
    return await db.roles.find_one({"id": role_id}, {"_id": 0})

@api_router.delete("/roles/{role_id}")
async def delete_role(role_id: str, request: Request, user: dict = Depends(get_current_user)):
    role = await db.roles.find_one({"id": role_id, "org_id": user['org_id']}, {"_id": 0})
    if not role:
        raise HTTPException(status_code=404, detail="Role not found")
    if role.get('is_system'):
        raise HTTPException(status_code=400, detail="Cannot delete system roles")
    
    await db.roles.delete_one({"id": role_id})
    await log_audit(user['org_id'], "role_deleted", "role", user['id'], user['email'], role_id,
                   {"name": role['name']}, request.client.host if request.client else None)
    return {"message": "Role deleted"}

@api_router.get("/permissions")
async def list_permissions(user: dict = Depends(get_current_user)):
    permissions = await db.permissions.find({}, {"_id": 0}).to_list(100)
    return permissions

# ===================== GROUP ROUTES =====================

@api_router.get("/groups")
async def list_groups(user: dict = Depends(get_current_user)):
    groups = await db.groups.find({"org_id": user['org_id']}, {"_id": 0}).to_list(100)
    # Add member count
    for group in groups:
        group['member_count'] = await db.users.count_documents({"group_ids": group['id']})
    return groups

@api_router.post("/groups")
async def create_group(group: GroupCreate, request: Request, user: dict = Depends(get_current_user)):
    if group.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    group_id = str(uuid.uuid4())
    group_doc = {
        "id": group_id,
        "name": group.name,
        "description": group.description,
        "org_id": group.org_id,
        "parent_id": group.parent_id,
        "role_ids": group.role_ids,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.groups.insert_one(group_doc)
    await log_audit(user['org_id'], "group_created", "group", user['id'], user['email'], group_id,
                   {"name": group.name}, request.client.host if request.client else None)
    return {**group_doc, "_id": None, "member_count": 0}

@api_router.put("/groups/{group_id}")
async def update_group(group_id: str, update: dict, request: Request, user: dict = Depends(get_current_user)):
    group = await db.groups.find_one({"id": group_id, "org_id": user['org_id']}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    update.pop('id', None)
    update.pop('org_id', None)
    await db.groups.update_one({"id": group_id}, {"$set": update})
    await log_audit(user['org_id'], "group_updated", "group", user['id'], user['email'], group_id,
                   update, request.client.host if request.client else None)
    return await db.groups.find_one({"id": group_id}, {"_id": 0})

@api_router.delete("/groups/{group_id}")
async def delete_group(group_id: str, request: Request, user: dict = Depends(get_current_user)):
    group = await db.groups.find_one({"id": group_id, "org_id": user['org_id']}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    # Remove group from users
    await db.users.update_many({"group_ids": group_id}, {"$pull": {"group_ids": group_id}})
    await db.groups.delete_one({"id": group_id})
    await log_audit(user['org_id'], "group_deleted", "group", user['id'], user['email'], group_id,
                   {"name": group['name']}, request.client.host if request.client else None)
    return {"message": "Group deleted"}

@api_router.post("/groups/{group_id}/members")
async def add_group_members(group_id: str, user_ids: List[str], request: Request, user: dict = Depends(get_current_user)):
    group = await db.groups.find_one({"id": group_id, "org_id": user['org_id']}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    await db.users.update_many(
        {"id": {"$in": user_ids}, "org_id": user['org_id']},
        {"$addToSet": {"group_ids": group_id}}
    )
    await log_audit(user['org_id'], "group_members_added", "group", user['id'], user['email'], group_id,
                   {"user_ids": user_ids}, request.client.host if request.client else None)
    return {"message": f"Added {len(user_ids)} members to group"}

@api_router.delete("/groups/{group_id}/members")
async def remove_group_members(group_id: str, user_ids: List[str], request: Request, user: dict = Depends(get_current_user)):
    group = await db.groups.find_one({"id": group_id, "org_id": user['org_id']}, {"_id": 0})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    
    await db.users.update_many(
        {"id": {"$in": user_ids}},
        {"$pull": {"group_ids": group_id}}
    )
    return {"message": f"Removed {len(user_ids)} members from group"}

# ===================== USER ROUTES =====================

@api_router.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    users = await db.users.find({"org_id": user['org_id']}, {"_id": 0, "password": 0}).to_list(1000)
    return users


@api_router.get("/users/export")
async def export_users(format: str = "xlsx", user: dict = Depends(get_current_user)):
    """Export all users as CSV or Excel with all HR fields"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can export users")

    users = await db.users.find(
        {"org_id": user["org_id"]}, {"_id": 0, "password": 0}
    ).to_list(5000)

    # Get app assignments
    saml_apps = await db.saml_apps.find({"org_id": user["org_id"]}, {"_id": 0, "name": 1, "approved_user_ids": 1}).to_list(100)
    oidc_apps = await db.oidc_apps.find({"org_id": user["org_id"]}, {"_id": 0, "name": 1, "approved_user_ids": 1}).to_list(100)
    all_apps = saml_apps + oidc_apps

    def get_user_apps(uid):
        return ", ".join(a["name"] for a in all_apps if uid in a.get("approved_user_ids", []))

    columns = [
        ("Employee ID", "adrenalin_employee_id"),
        ("Title", "title"),
        ("First Name", "first_name"),
        ("Last Name", "last_name"),
        ("Full Name", "name"),
        ("Email", "email"),
        ("Personal Email", "personal_email"),
        ("Work Mobile", "work_mobile"),
        ("Personal Mobile", "employee_mobile"),
        ("Gender", "sex"),
        ("Date of Birth", "date_of_birth"),
        ("PAN Number", "pan_number"),
        ("Designation", "designation"),
        ("Department", "department"),
        ("Department Code", "department_code"),
        ("Grade", "grade"),
        ("Company", "company"),
        ("Legal Entity", "legal_entity_code"),
        ("Business Line", "business_line"),
        ("Branch", "branch_code"),
        ("Location", "location"),
        ("Office Location", "office_location"),
        ("Pincode", "employee_pincode"),
        ("Employee Status", "employee_status"),
        ("Employee Status Desc", "employee_status_description"),
        ("Employment Status", "employment_status"),
        ("Employment Status Desc", "employment_status_description"),
        ("Joining Date", "joining_date"),
        ("Date of Exit", "date_of_exit"),
        ("Added On", "emp_added_on"),
        ("L1 Manager Name", "supervisor_name"),
        ("L1 Manager Email", "supervisor_email"),
        ("L1 Manager Code", "supervisor_employee_code"),
        ("L2 Manager Name", "l2_manager_name"),
        ("L2 Manager Email", "l2_manager_email"),
        ("L2 Manager Code", "l2_manager_employee_code"),
        ("System Role", "role"),
        ("Status", "status"),
        ("Created Via", "created_via"),
        ("Last HR Sync", "hr_synced_at"),
        ("Assigned Apps", None),
    ]

    import io
    if format == "csv":
        import csv
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow([c[0] for c in columns])
        for u in users:
            row = []
            for label, key in columns:
                if key is None:
                    row.append(get_user_apps(u.get("id", "")))
                else:
                    row.append(str(u.get(key, "") or ""))
            writer.writerow(row)
        content = output.getvalue().encode("utf-8-sig")
        from fastapi.responses import Response
        return Response(
            content=content,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=users_export.csv"},
        )
    else:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"

        # Header styling
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
        thin_border = Border(
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Write headers
        for col_idx, (label, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=label)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, u in enumerate(users, 2):
            for col_idx, (label, key) in enumerate(columns, 1):
                if key is None:
                    val = get_user_apps(u.get("id", ""))
                else:
                    val = str(u.get(key, "") or "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center")

        # Auto-width columns
        for col_idx in range(1, len(columns) + 1):
            max_len = max(
                len(str(ws.cell(row=r, column=col_idx).value or ""))
                for r in range(1, min(len(users) + 2, 50))
            )
            ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 4, 35)

        # Freeze header row
        ws.freeze_panes = "A2"

        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        from fastapi.responses import Response
        return Response(
            content=content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=users_export.xlsx"},
        )

@api_router.post("/users")
async def create_user(new_user: UserCreate, request: Request, user: dict = Depends(get_current_user)):
    if new_user.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    existing = await db.users.find_one({"email": new_user.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already exists")
    
    user_id = str(uuid.uuid4())
    user_doc = {
        "id": user_id,
        "email": new_user.email,
        "password": hash_password(new_user.password),
        "name": new_user.name,
        "org_id": new_user.org_id,
        "role": "user",
        "group_ids": [],
        "role_ids": [],
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.users.insert_one(user_doc)
    await log_audit(user['org_id'], "user_created", "user", user['id'], user['email'], user_id,
                   {"name": new_user.name, "email": new_user.email}, request.client.host if request.client else None)
    
    return {k: v for k, v in user_doc.items() if k != 'password' and k != '_id'}

@api_router.put("/users/{user_id}")
async def update_user(user_id: str, update: UserUpdate, request: Request, user: dict = Depends(get_current_user)):
    target_user = await db.users.find_one({"id": user_id, "org_id": user['org_id']}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_data = {k: v for k, v in update.model_dump().items() if v is not None}
    if update_data:
        await db.users.update_one({"id": user_id}, {"$set": update_data})
        await log_audit(user['org_id'], "user_updated", "user", user['id'], user['email'], user_id,
                       update_data, request.client.host if request.client else None)
        # Push updated user to Kissflow in background
        try:
            email = target_user.get("email", "")
            if email:
                await push_single_user_to_kissflow(db, user["org_id"], email)
        except Exception as e:
            logging.getLogger("kissflow_scim").warning(f"Kissflow push after user update failed for {email}: {e}")
    
    return await db.users.find_one({"id": user_id}, {"_id": 0, "password": 0})


@api_router.post("/users/{user_id}/reset-password")
async def reset_user_password(user_id: str, body: dict, request: Request, user: dict = Depends(get_current_user)):
    """Admin sets a custom password for a user"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can reset passwords")

    new_password = body.get("password", "")
    if len(new_password) < 6:
        raise HTTPException(status_code=400, detail="Password must be at least 6 characters")

    target = await db.users.find_one({"id": user_id, "org_id": user["org_id"]}, {"_id": 0})
    if not target:
        raise HTTPException(status_code=404, detail="User not found")

    hashed = hash_password(new_password)
    await db.users.update_one({"id": user_id}, {"$set": {"password": hashed}})
    await log_audit(user["org_id"], "password_reset", "user", user["id"], user["email"], user_id,
                    {"target_email": target["email"]}, request.client.host if request.client else None)

    return {"message": f"Password reset for {target['email']}"}


@api_router.delete("/users/{user_id}")
async def delete_user(user_id: str, request: Request, user: dict = Depends(get_current_user)):
    if user_id == user['id']:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    
    target_user = await db.users.find_one({"id": user_id, "org_id": user['org_id']}, {"_id": 0})
    if not target_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    await db.users.delete_one({"id": user_id})
    await log_audit(user['org_id'], "user_deleted", "user", user['id'], user['email'], user_id,
                   {"email": target_user['email']}, request.client.host if request.client else None)
    return {"message": "User deleted"}

# ===================== SAML APP ROUTES =====================

@api_router.get("/apps/saml")
async def list_saml_apps(user: dict = Depends(get_current_user)):
    apps = await db.saml_apps.find({"org_id": user['org_id']}, {"_id": 0, "private_key": 0}).to_list(100)
    return apps

@api_router.post("/apps/saml")
async def create_saml_app(app: SAMLAppCreate, request: Request, user: dict = Depends(get_current_user)):
    if app.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    cert, key = generate_self_signed_cert()
    app_id = str(uuid.uuid4())
    
    app_doc = {
        "id": app_id,
        "name": app.name,
        "description": app.description,
        "org_id": app.org_id,
        "entity_id": app.entity_id,
        "acs_url": app.acs_url,
        "slo_url": app.slo_url,
        "home_url": app.home_url,
        "name_id_format": app.name_id_format,
        "sign_assertions": app.sign_assertions,
        "sign_response": app.sign_response,
        "attribute_mappings": app.attribute_mappings,
        "certificate": cert,
        "private_key": key,
        "logo_url": app.logo_url,
        "allowed_group_ids": app.allowed_group_ids,
        "allowed_role_ids": app.allowed_role_ids,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.saml_apps.insert_one(app_doc)
    await log_audit(user['org_id'], "saml_app_created", "app", user['id'], user['email'], app_id,
                   {"name": app.name}, request.client.host if request.client else None)
    
    return {k: v for k, v in app_doc.items() if k != 'private_key' and k != '_id'}

@api_router.get("/apps/saml/{app_id}")
async def get_saml_app(app_id: str, user: dict = Depends(get_current_user)):
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0, "private_key": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    return app

@api_router.put("/apps/saml/{app_id}")
async def update_saml_app(app_id: str, update: dict, request: Request, user: dict = Depends(get_current_user)):
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    update.pop('id', None)
    update.pop('org_id', None)
    update.pop('certificate', None)
    update.pop('private_key', None)
    
    await db.saml_apps.update_one({"id": app_id}, {"$set": update})
    await log_audit(user['org_id'], "saml_app_updated", "app", user['id'], user['email'], app_id,
                   update, request.client.host if request.client else None)
    return await db.saml_apps.find_one({"id": app_id}, {"_id": 0, "private_key": 0})

@api_router.delete("/apps/saml/{app_id}")
async def delete_saml_app(app_id: str, request: Request, user: dict = Depends(get_current_user)):
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    await db.saml_apps.delete_one({"id": app_id})
    await log_audit(user['org_id'], "saml_app_deleted", "app", user['id'], user['email'], app_id,
                   {"name": app['name']}, request.client.host if request.client else None)
    return {"message": "App deleted"}


@api_router.get("/apps/saml/{app_id}/users")
async def get_saml_app_users(app_id: str, user: dict = Depends(get_current_user)):
    """Get users assigned to a SAML app"""
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    approved_ids = app.get('approved_user_ids', [])
    users = []
    for uid in approved_ids:
        u = await db.users.find_one({"id": uid}, {"_id": 0, "password": 0})
        if u:
            users.append({"id": u['id'], "email": u.get('email'), "name": u.get('name'), "role": u.get('role')})
    return users

@api_router.post("/apps/saml/{app_id}/users")
async def assign_users_to_saml_app(app_id: str, body: dict, request: Request, user: dict = Depends(get_current_user)):
    """Assign users to a SAML app"""
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    user_ids = body.get('user_ids', [])
    await db.saml_apps.update_one({"id": app_id}, {"$addToSet": {"approved_user_ids": {"$each": user_ids}}})
    return {"message": f"Added {len(user_ids)} user(s)", "user_ids": user_ids}

@api_router.delete("/apps/saml/{app_id}/users/{user_id}")
async def remove_user_from_saml_app(app_id: str, user_id: str, request: Request, user: dict = Depends(get_current_user)):
    """Remove a user from a SAML app"""
    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    await db.saml_apps.update_one({"id": app_id}, {"$pull": {"approved_user_ids": user_id}})
    return {"message": "User removed"}


@api_router.get("/apps/saml/{app_id}/metadata")
async def get_saml_metadata(app_id: str, request: Request):
    app = await db.saml_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    base_url = get_public_base_url(request)
    metadata = generate_saml_metadata(app, base_url)
    return Response(content=metadata, media_type="application/xml")

@api_router.get("/apps/saml/{app_id}/kissflow-config")
async def get_kissflow_config(app_id: str, request: Request):
    """Get configuration values for Kissflow SSO setup (IdP URL, Security Key, etc.)"""
    import hashlib
    import base64
    
    app = await db.saml_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    base_url = get_public_base_url(request)
    
    # Extract certificate and calculate SHA256 fingerprint
    cert_pem = app.get('certificate', '')
    fingerprint = ''
    
    if cert_pem:
        cert_b64 = cert_pem.replace('-----BEGIN CERTIFICATE-----', '').replace('-----END CERTIFICATE-----', '').replace('\n', '').strip()
        try:
            cert_der = base64.b64decode(cert_b64)
            fingerprint = hashlib.sha256(cert_der).hexdigest()
        except:
            fingerprint = 'Error calculating fingerprint'
    
    return {
        "app_name": app.get('name'),
        "entity_id": app.get('entity_id'),
        "idp_url": f"{base_url}/api/saml/{app_id}/sso",
        "sso_url": f"{base_url}/api/saml/{app_id}/sso",
        "slo_url": f"{base_url}/api/saml/{app_id}/slo",
        "security_key": fingerprint,
        "security_key_formatted": ':'.join(fingerprint[i:i+2] for i in range(0, len(fingerprint), 2)).upper() if fingerprint else '',
        "metadata_url": f"{base_url}/api/saml/{app_id}/metadata",
        "certificate": cert_pem,
        "name_id_format": app.get('name_id_format', 'urn:oasis:names:tc:SAML:1.1:nameid-format:emailAddress'),
        "instructions": {
            "step1": "In Kissflow Admin > SSO Settings, select Manual configuration",
            "step2": f"Set IdP URL to: {base_url}/api/saml/{app_id}/sso",
            "step3": f"Set Sign-out URL to: {base_url}/api/saml/{app_id}/slo (optional)",
            "step4": f"Set Security Key to: {fingerprint}",
            "step5": "Save and test the SSO connection"
        }
    }

@api_router.get("/saml/{app_id}/sso")
@api_router.post("/saml/{app_id}/sso")
async def saml_sso(app_id: str, request: Request):
    """SAML Single Sign-On endpoint - handles both IdP-initiated and SP-initiated SSO"""
    app = await db.saml_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="SAML App not found")
    
    base_url = get_public_base_url(request)
    frontend_url = base_url
    
    # Capture RelayState if present (SP-initiated flow)
    params = dict(request.query_params)
    relay_state = params.get('RelayState', '')
    
    # Build login URL with SSO app and optional relay state
    login_url = f"{frontend_url}/login?sso_app={app_id}"
    if relay_state:
        from urllib.parse import quote
        login_url += f"&relay_state={quote(relay_state)}"

    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>SSO Login - {app.get('name', 'SAML App')}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'IBM Plex Sans', -apple-system, sans-serif; background: #FAFAFA; min-height: 100vh; display: flex; align-items: center; justify-content: center; }}
        .card {{ background: white; border: 1px solid #e5e5e5; padding: 48px; max-width: 480px; width: 90%; }}
        h1 {{ font-size: 24px; font-weight: 800; margin-bottom: 8px; }}
        p {{ color: #71717a; margin-bottom: 24px; }}
        .app-info {{ background: #f4f4f5; padding: 16px; margin-bottom: 24px; }}
        .app-name {{ font-weight: 600; font-size: 18px; }}
        .app-url {{ font-size: 12px; color: #71717a; font-family: monospace; word-break: break-all; }}
        .btn {{ display: block; width: 100%; padding: 16px; background: #0051FF; color: white; text-align: center; text-decoration: none; font-weight: 600; margin-bottom: 12px; }}
        .btn:hover {{ background: #003ECC; }}
        .btn-secondary {{ background: #f4f4f5; color: #18181b; }}
        .btn-secondary:hover {{ background: #e5e5e5; }}
        .info {{ font-size: 12px; color: #a1a1aa; margin-top: 24px; text-align: center; }}
    </style>
</head>
<body>
    <div class="card">
        <h1>Single Sign-On</h1>
        <p>You are about to sign in to:</p>
        
        <div class="app-info">
            <div class="app-name">{app.get('name', 'Application')}</div>
            <div class="app-url">{app.get('entity_id', '')}</div>
        </div>
        
        <a href="{login_url}" class="btn">Sign In with Kissflow IAM</a>
        <a href="{app.get('acs_url', '#')}" class="btn btn-secondary">Cancel</a>
        
        <p class="info">
            This is the SAML Identity Provider endpoint.<br>
            Entity ID: {app.get('entity_id')}<br>
            ACS URL: {app.get('acs_url')}
        </p>
    </div>
</body>
</html>'''
    
    return Response(content=html_content, media_type="text/html")

@api_router.get("/saml/{app_id}/slo")
@api_router.post("/saml/{app_id}/slo")
async def saml_slo(app_id: str, request: Request):
    """SAML Single Logout endpoint"""
    app = await db.saml_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="SAML App not found")
    
    # For SLO, redirect back to the app's SLO URL or a logout confirmation page
    slo_url = app.get('slo_url')
    if slo_url:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=slo_url)
    
    html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Logged Out - {app.get('name', 'SAML App')}</title>
    <style>
        * {{ margin: 0; padding: 0; box-sizing: border-box; }}
        body {{ font-family: 'IBM Plex Sans', -apple-system, sans-serif; background: #FAFAFA; min-height: 100vh; display: flex; align-items: center; justify-content: center; }}
        .card {{ background: white; border: 1px solid #e5e5e5; padding: 48px; max-width: 400px; text-align: center; }}
        h1 {{ font-size: 24px; font-weight: 800; margin-bottom: 16px; color: #00CC66; }}
        p {{ color: #71717a; }}
    </style>
</head>
<body>
    <div class="card">
        <h1>Successfully Logged Out</h1>
        <p>You have been logged out of {app.get('name', 'the application')}.</p>
    </div>
</body>
</html>'''
    
    return Response(content=html_content, media_type="text/html")

@api_router.get("/saml/{app_id}/complete")
async def saml_complete_sso(app_id: str, request: Request, token: str = None, relay_state: str = None, debug: int = 0):
    """Complete SAML SSO - Generate signed SAML Response and POST to ACS URL"""
    import base64
    import uuid as uuid_module
    from xml.sax.saxutils import escape
    from lxml import etree
    
    app = await db.saml_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="SAML App not found")
    
    base_url = get_public_base_url(request)
    
    # Get token from query param or Authorization header
    auth_token = token
    if not auth_token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            auth_token = auth_header[7:]
    
    if not auth_token:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"{base_url}/login?sso_app={app_id}")
    
    # Decode token and get user
    try:
        payload = decode_token(auth_token)
    except:
        from fastapi.responses import RedirectResponse
        return RedirectResponse(url=f"{base_url}/login?sso_app={app_id}")
    
    user = await db.users.find_one({"id": payload['user_id']}, {"_id": 0, "password": 0})
    if not user:
        raise HTTPException(status_code=401, detail="User not found")
    
    # Check user has access to this app
    has_access = await check_user_app_access(user, app)
    
    # For SP-initiated SSO: if user doesn't have direct access to this app,
    # check if they have access to any sibling app sharing the same ACS URL.
    # Kissflow (SP) only knows about one SAML connection, but we may have 
    # multiple sub-apps (Expense Mgmt, Travel Mgmt, etc.) pointing to the same SP.
    if not has_access:
        acs_url = app.get('acs_url', '')
        if acs_url:
            sibling_apps = await db.saml_apps.find(
                {"org_id": user.get('org_id'), "acs_url": acs_url, "id": {"$ne": app_id}, "status": "active"},
                {"_id": 0}
            ).to_list(100)
            for sibling in sibling_apps:
                if await check_user_app_access(user, sibling):
                    has_access = True
                    break
    
    if not has_access:
        raise HTTPException(status_code=403, detail="You don't have access to this application")
    
    # Generate SAML Response timestamps
    now = datetime.now(timezone.utc)
    not_on_or_after = now + timedelta(minutes=5)
    response_id = f"_{''.join(str(uuid_module.uuid4()).split('-'))}"
    assertion_id = f"_{''.join(str(uuid_module.uuid4()).split('-'))}"
    
    acs_url = app.get('acs_url', '')
    name_id = user.get('email', '')
    name_id_format = app.get('name_id_format', 'urn:oasis:names:tc:SAML:1.1:nameid-format:emailAddress')
    
    # Determine issuer, cert, key - if this app shares entity_id + acs_url with another app,
    # use the original (parent) app's issuer/cert/key so the SP can validate our response
    issuer = f"{base_url}/api/saml/{app_id}"
    cert_pem = app.get('certificate', '')
    key_pem = app.get('private_key', '')
    
    # Find the FIRST (earliest) app with same entity_id + acs_url (the "primary" SAML connection)
    primary_app = await db.saml_apps.find_one({
        "entity_id": app.get('entity_id'),
        "acs_url": acs_url,
        "org_id": app.get('org_id')
    }, {"_id": 0, "id": 1, "certificate": 1, "private_key": 1}, sort=[("created_at", 1)])
    
    if primary_app and primary_app['id'] != app_id:
        issuer = f"{base_url}/api/saml/{primary_app['id']}"
        if primary_app.get('certificate'):
            cert_pem = primary_app['certificate']
        if primary_app.get('private_key'):
            key_pem = primary_app['private_key']
        logging.info(f"App {app_id} shares SP with primary {primary_app['id']}, using primary issuer")
    
    now_str = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    not_on_or_after_str = not_on_or_after.strftime('%Y-%m-%dT%H:%M:%SZ')
    
    # Define SAML namespaces
    NSMAP = {
        'samlp': 'urn:oasis:names:tc:SAML:2.0:protocol',
        'saml': 'urn:oasis:names:tc:SAML:2.0:assertion',
        'xs': 'http://www.w3.org/2001/XMLSchema',
        'xsi': 'http://www.w3.org/2001/XMLSchema-instance',
    }
    
    XS_NS = 'http://www.w3.org/2001/XMLSchema'
    XSI_NS = 'http://www.w3.org/2001/XMLSchema-instance'
    
    # Build SAML Response XML using lxml
    response_elem = etree.Element('{urn:oasis:names:tc:SAML:2.0:protocol}Response', nsmap=NSMAP)
    response_elem.set('ID', response_id)
    response_elem.set('Version', '2.0')
    response_elem.set('IssueInstant', now_str)
    response_elem.set('Destination', acs_url)
    
    # Issuer
    issuer_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Issuer')
    issuer_elem.text = issuer
    
    # Status
    status_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:protocol}Status')
    status_code_elem = etree.SubElement(status_elem, '{urn:oasis:names:tc:SAML:2.0:protocol}StatusCode')
    status_code_elem.set('Value', 'urn:oasis:names:tc:SAML:2.0:status:Success')
    
    # Assertion
    assertion_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Assertion')
    assertion_elem.set('ID', assertion_id)
    assertion_elem.set('Version', '2.0')
    assertion_elem.set('IssueInstant', now_str)
    assertion_elem.set('{http://www.w3.org/2001/XMLSchema-instance}schemaLocation', 
                       'urn:oasis:names:tc:SAML:2.0:assertion http://docs.oasis-open.org/security/saml/v2.0/saml-schema-assertion-2.0.xsd')
    
    # Assertion > Issuer
    assertion_issuer = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Issuer')
    assertion_issuer.text = issuer
    
    # Assertion > Subject
    subject_elem = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Subject')
    name_id_elem = etree.SubElement(subject_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}NameID')
    name_id_elem.set('Format', name_id_format)
    name_id_elem.text = name_id
    
    subj_conf = etree.SubElement(subject_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}SubjectConfirmation')
    subj_conf.set('Method', 'urn:oasis:names:tc:SAML:2.0:cm:bearer')
    subj_conf_data = etree.SubElement(subj_conf, '{urn:oasis:names:tc:SAML:2.0:assertion}SubjectConfirmationData')
    subj_conf_data.set('NotOnOrAfter', not_on_or_after_str)
    subj_conf_data.set('Recipient', acs_url)
    
    # Assertion > Conditions
    conditions_elem = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Conditions')
    conditions_elem.set('NotBefore', now_str)
    conditions_elem.set('NotOnOrAfter', not_on_or_after_str)
    audience_restriction = etree.SubElement(conditions_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AudienceRestriction')
    audience_elem = etree.SubElement(audience_restriction, '{urn:oasis:names:tc:SAML:2.0:assertion}Audience')
    audience_elem.text = app.get('entity_id', '')
    
    # Assertion > AuthnStatement
    authn_stmt = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnStatement')
    authn_stmt.set('AuthnInstant', now_str)
    authn_ctx = etree.SubElement(authn_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnContext')
    authn_ctx_ref = etree.SubElement(authn_ctx, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnContextClassRef')
    authn_ctx_ref.text = 'urn:oasis:names:tc:SAML:2.0:ac:classes:Password'
    
    # Assertion > AttributeStatement
    attr_stmt = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeStatement')
    
    email_attr = etree.SubElement(attr_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}Attribute')
    email_attr.set('Name', 'email')
    email_attr.set('NameFormat', 'urn:oasis:names:tc:SAML:2.0:attrname-format:basic')
    email_val = etree.SubElement(email_attr, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeValue')
    email_val.set(f'{{{XSI_NS}}}type', 'xs:string')
    email_val.text = user.get('email', '')
    
    name_attr = etree.SubElement(attr_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}Attribute')
    name_attr.set('Name', 'name')
    name_attr.set('NameFormat', 'urn:oasis:names:tc:SAML:2.0:attrname-format:basic')
    name_val = etree.SubElement(name_attr, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeValue')
    name_val.set(f'{{{XSI_NS}}}type', 'xs:string')
    name_val.text = user.get('name', '')
    
    # Sign the SAML Response (cert_pem and key_pem already set above, potentially from parent app)
    signed_response_xml = None
    
    if key_pem and cert_pem:
        try:
            import signxml
            from signxml import XMLSigner
            
            # Sign the assertion using signxml (enveloped signature)
            signer = XMLSigner(
                method=signxml.methods.enveloped,
                signature_algorithm="rsa-sha256",
                digest_algorithm="sha256",
                c14n_algorithm="http://www.w3.org/2001/10/xml-exc-c14n#"
            )
            
            # Sign the assertion element
            signed_assertion = signer.sign(
                assertion_elem, 
                key=key_pem, 
                cert=cert_pem, 
                reference_uri=f'#{assertion_id}'
            )
            
            # Replace the unsigned assertion with the signed one in the response
            # Find and remove old assertion, add signed one
            for child in list(response_elem):
                if child.tag == '{urn:oasis:names:tc:SAML:2.0:assertion}Assertion':
                    response_elem.remove(child)
            response_elem.append(signed_assertion)
            
            # Move Signature to correct position per SAML schema: right after Issuer (index 1)
            # signxml appends Signature at the end; SAML requires Issuer, Signature, Subject, ...
            ds_ns = 'http://www.w3.org/2000/09/xmldsig#'
            sig_elem = signed_assertion.find(f'{{{ds_ns}}}Signature')
            if sig_elem is not None:
                signed_assertion.remove(sig_elem)
                # Insert after Issuer (position 1)
                signed_assertion.insert(1, sig_elem)
            
            # Clean base64 text content in signature elements (remove PEM newlines)
            # Kissflow's strict parser rejects newlines inside X509Certificate, SignatureValue, DigestValue
            for tag in ['X509Certificate', 'SignatureValue', 'DigestValue']:
                for elem in response_elem.iter(f'{{{ds_ns}}}{tag}'):
                    if elem.text:
                        elem.text = elem.text.replace('\n', '').replace('\r', '').replace(' ', '')
            
            signed_response_xml = etree.tostring(response_elem, xml_declaration=False, encoding='unicode', pretty_print=False)
            logging.info(f"SAML Response signed successfully for app {app_id}, user {name_id}")
        except Exception as e:
            logging.error(f"SAML signing failed: {e}")
            signed_response_xml = None
    
    if not signed_response_xml:
        signed_response_xml = etree.tostring(response_elem, xml_declaration=False, encoding='unicode', pretty_print=False)
        logging.warning(f"Sending UNSIGNED SAML response for app {app_id}")

    # Base64 encode the SAML Response - ensure clean encoding
    xml_clean = signed_response_xml.strip()
    saml_response_b64 = base64.b64encode(xml_clean.encode('utf-8')).decode('ascii')
    
    # Log the SSO attempt
    await log_audit(user['org_id'], "saml_sso_initiated", "app", user['id'], user['email'], app_id,
                   {"app_name": app.get('name'), "acs_url": acs_url}, 
                   request.client.host if request.client else None)
    
    # Return an HTML form that auto-submits to the ACS URL
    # Log the base64 length for debugging
    logging.info(f"SAML base64 length: {len(saml_response_b64)}, mod4: {len(saml_response_b64) % 4}")
    
    if debug:
        # Debug mode: show diagnostic info and submit to our debug endpoint first
        debug_url = f"{base_url}/api/saml/debug/receive"
        html_content = f'''<!DOCTYPE html>
<html lang="en">
<head><meta charset="UTF-8"><title>SAML Debug Mode</title></head>
<body style="font-family:sans-serif;padding:20px;max-width:800px;margin:0 auto">
<h2>SAML SSO Debug Mode</h2>
<div id="diagnostics" style="background:#f5f5f5;padding:15px;border-radius:8px;margin:15px 0;font-family:monospace;font-size:13px"></div>
<form id="debugForm" method="POST" action="{debug_url}">
    <input type="hidden" name="SAMLResponse" id="samlResponse"/>
    <input type="hidden" name="ACS" value="{escape(acs_url)}"/>
    {'<input type="hidden" name="RelayState" value="' + escape(relay_state) + '"/>' if relay_state else ''}
    <button type="submit" style="padding:12px 24px;font-size:16px;background:#10b981;color:white;border:none;border-radius:6px;cursor:pointer;margin:5px">
        Step 1: Verify data via debug endpoint
    </button>
</form>
<form id="directForm" method="POST" action="{escape(acs_url)}">
    <input type="hidden" name="SAMLResponse" id="samlResponseDirect"/>
    {'<input type="hidden" name="RelayState" value="' + escape(relay_state) + '"/>' if relay_state else ''}
    <button type="submit" style="padding:12px 24px;font-size:16px;background:#3b82f6;color:white;border:none;border-radius:6px;cursor:pointer;margin:5px">
        Step 2: Submit directly to Kissflow
    </button>
</form>
<script>
var samlData = "{saml_response_b64}";
document.getElementById('samlResponse').value = samlData;
document.getElementById('samlResponseDirect').value = samlData;
var diag = document.getElementById('diagnostics');
var lines = [];
lines.push("Base64 length: " + samlData.length);
lines.push("Length mod 4: " + (samlData.length % 4));
lines.push("Has newlines: " + (samlData.indexOf("\\n") >= 0));
lines.push("Has spaces: " + (samlData.indexOf(" ") >= 0));
lines.push("Starts with: " + samlData.substring(0, 60) + "...");
lines.push("Ends with: ..." + samlData.substring(samlData.length - 40));
try {{
    var decoded = atob(samlData);
    lines.push("Browser atob() decode: SUCCESS (" + decoded.length + " chars)");
    lines.push("XML starts with: " + decoded.substring(0, 80));
}} catch(e) {{
    lines.push("Browser atob() decode: FAILED - " + e.message);
}}
diag.innerHTML = lines.join("<br>");
</script>
</body></html>'''
    else:
        # Kissflow's SAML parser cannot handle RelayState (it tries to base64/JSON decode it).
        # Omit RelayState entirely.
        # If the app has a home_url, submit SAML via iframe to authenticate,
        # then redirect browser to the specific module URL.
        home_url = app.get('home_url', '')
        
        if home_url:
            # Iframe-based auth + redirect to specific module
            # Include RelayState if present (critical for SP-initiated SSO, especially mobile)
            relay_field = f'<input type="hidden" name="RelayState" value="{escape(relay_state)}"/>' if relay_state else ''
            html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Signing in to {escape(app.get('name', 'Application'))}...</title>
    <style>
        body {{ font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif; display: flex; align-items: center; justify-content: center; min-height: 100vh; margin: 0; background: #f9fafb; }}
        .loader {{ text-align: center; }}
        .spinner {{ width: 36px; height: 36px; border: 3px solid #e2e8f0; border-top-color: #10b981; border-radius: 50%; animation: spin 0.8s linear infinite; margin: 0 auto 16px; }}
        @keyframes spin {{ to {{ transform: rotate(360deg); }} }}
        p {{ color: #64748b; font-size: 14px; }}
    </style>
</head>
<body>
    <div class="loader">
        <div class="spinner"></div>
        <p>Signing in to {escape(app.get('name', 'Application'))}...</p>
    </div>
    <iframe name="authFrame" style="display:none"></iframe>
    <form id="samlForm" method="POST" action="{escape(acs_url)}" target="authFrame">
        <input type="hidden" name="SAMLResponse" id="samlResponse"/>
        {relay_field}
    </form>
    <script>
        var samlData = "{saml_response_b64}";
        var moduleUrl = "{escape(home_url)}";
        var hasRelayState = {"true" if relay_state else "false"};
        document.getElementById('samlResponse').value = samlData;
        if (hasRelayState) {{
            // SP-initiated flow (e.g., from Kissflow native app): submit directly
            // so Kissflow can process RelayState and redirect back to native app
            document.getElementById('samlForm').removeAttribute('target');
            document.getElementById('samlForm').submit();
        }} else {{
            // IdP-initiated flow: use iframe auth + redirect to module
            document.getElementById('samlForm').submit();
            setTimeout(function() {{
                window.location.href = moduleUrl;
            }}, 2500);
            // If the module URL opens a native app via deep link,
            // the browser stays on this page. Redirect back to launcher.
            setTimeout(function() {{
                window.location.href = "/launcher";
            }}, 4000);
        }}
    </script>
</body>
</html>'''
        else:
            # No home_url: direct SAML form submit (lands on Kissflow homepage)
            # Include RelayState if present (critical for SP-initiated SSO / mobile return)
            relay_field = f'<input type="hidden" name="RelayState" value="{escape(relay_state)}"/>' if relay_state else ''
            html_content = f'''<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <title>Redirecting to {escape(app.get('name', 'Application'))}...</title>
</head>
<body>
    <form id="samlForm" method="POST" action="{escape(acs_url)}">
        <input type="hidden" name="SAMLResponse" id="samlResponse"/>
        {relay_field}
        <noscript>
            <input type="submit" value="Continue to {escape(app.get('name', 'Application'))}"/>
        </noscript>
    </form>
    <p style="font-family: sans-serif; color: #666; text-align: center; margin-top: 50px;">
        Redirecting to {escape(app.get('name', 'Application'))}...
    </p>
    <script>
        var samlData = "{saml_response_b64}";
        document.getElementById('samlResponse').value = samlData;
        document.getElementById('samlForm').submit();
    </script>
</body>
</html>'''
    
    return Response(content=html_content, media_type="text/html")



@api_router.post("/saml/debug/receive")
async def saml_debug_receive(request: Request):
    """Debug endpoint: receives SAML form POST, validates base64, re-submits to Kissflow"""
    import base64
    form = await request.form()
    saml_response = form.get('SAMLResponse', '')
    relay_state = form.get('RelayState', '')
    acs_url = form.get('ACS', '')
    
    # Validate the received SAMLResponse
    diagnostics = []
    diagnostics.append(f"Received SAMLResponse length: {len(saml_response)}")
    diagnostics.append(f"Length mod 4: {len(saml_response) % 4}")
    diagnostics.append(f"Has newlines: {chr(10) in saml_response}")
    diagnostics.append(f"Has spaces: {' ' in saml_response}")
    
    valid = False
    try:
        decoded = base64.b64decode(saml_response, validate=True)
        diagnostics.append(f"Strict base64 decode: SUCCESS ({len(decoded)} bytes)")
        diagnostics.append(f"XML starts with: {decoded[:80].decode('utf-8', errors='replace')}")
        valid = True
    except Exception as e:
        diagnostics.append(f"Strict base64 decode: FAILED - {e}")
    
    diag_html = '<br>'.join(diagnostics)
    
    html = f'''<!DOCTYPE html><html><head><meta charset="UTF-8"><title>SAML Debug</title></head>
<body style="font-family:monospace;padding:20px">
<h2>SAML Debug - Browser Received Data</h2>
<div style="background:#f0f0f0;padding:10px;margin:10px 0">{diag_html}</div>
<p>Base64 valid: <b>{"YES" if valid else "NO"}</b></p>
<p>First 100 chars: <code>{saml_response[:100]}</code></p>
<p>Last 50 chars: <code>{saml_response[-50:]}</code></p>
<hr>
<p>Click below to forward this EXACT data to Kissflow:</p>
<form method="POST" action="{acs_url}">
    <input type="hidden" name="SAMLResponse" value="{saml_response}"/>
    {'<input type="hidden" name="RelayState" value="' + relay_state + '"/>' if relay_state else ''}
    <button type="submit" style="padding:10px 20px;font-size:16px;cursor:pointer">Submit to Kissflow</button>
</form>
</body></html>'''
    return Response(content=html, media_type="text/html")



@api_router.get("/saml/{app_id}/test")
async def saml_test_sso(app_id: str, request: Request, user: dict = Depends(get_current_user)):
    """Test SAML SSO - Returns decoded SAML assertion details as JSON for inspection"""
    import base64
    import uuid as uuid_module
    from lxml import etree

    app = await db.saml_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="SAML App not found")

    base_url = get_public_base_url(request)

    now = datetime.now(timezone.utc)
    not_on_or_after = now + timedelta(minutes=5)
    response_id = f"_{''.join(str(uuid_module.uuid4()).split('-'))}"
    assertion_id = f"_{''.join(str(uuid_module.uuid4()).split('-'))}"

    issuer = f"{base_url}/api/saml/{app_id}"
    acs_url = app.get('acs_url', '')
    name_id = user.get('email', '')
    name_id_format = app.get('name_id_format', 'urn:oasis:names:tc:SAML:1.1:nameid-format:emailAddress')
    
    # If this app shares entity_id + acs_url with another app, use primary's issuer/cert/key
    cert_pem = app.get('certificate', '')
    key_pem = app.get('private_key', '')
    
    primary_app = await db.saml_apps.find_one({
        "entity_id": app.get('entity_id'),
        "acs_url": acs_url,
        "org_id": user['org_id']
    }, {"_id": 0, "id": 1, "certificate": 1, "private_key": 1}, sort=[("created_at", 1)])
    
    if primary_app and primary_app['id'] != app_id:
        issuer = f"{base_url}/api/saml/{primary_app['id']}"
        if primary_app.get('certificate'):
            cert_pem = primary_app['certificate']
        if primary_app.get('private_key'):
            key_pem = primary_app['private_key']
    
    now_str = now.strftime('%Y-%m-%dT%H:%M:%SZ')
    not_on_or_after_str = not_on_or_after.strftime('%Y-%m-%dT%H:%M:%SZ')

    NSMAP = {
        'samlp': 'urn:oasis:names:tc:SAML:2.0:protocol',
        'saml': 'urn:oasis:names:tc:SAML:2.0:assertion',
    }

    response_elem = etree.Element('{urn:oasis:names:tc:SAML:2.0:protocol}Response', nsmap=NSMAP)
    response_elem.set('ID', response_id)
    response_elem.set('Version', '2.0')
    response_elem.set('IssueInstant', now_str)
    response_elem.set('Destination', acs_url)

    issuer_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Issuer')
    issuer_elem.text = issuer

    status_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:protocol}Status')
    status_code_elem = etree.SubElement(status_elem, '{urn:oasis:names:tc:SAML:2.0:protocol}StatusCode')
    status_code_elem.set('Value', 'urn:oasis:names:tc:SAML:2.0:status:Success')

    assertion_elem = etree.SubElement(response_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Assertion')
    assertion_elem.set('ID', assertion_id)
    assertion_elem.set('Version', '2.0')
    assertion_elem.set('IssueInstant', now_str)

    assertion_issuer = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Issuer')
    assertion_issuer.text = issuer

    subject_elem = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Subject')
    name_id_elem = etree.SubElement(subject_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}NameID')
    name_id_elem.set('Format', name_id_format)
    name_id_elem.text = name_id

    subj_conf = etree.SubElement(subject_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}SubjectConfirmation')
    subj_conf.set('Method', 'urn:oasis:names:tc:SAML:2.0:cm:bearer')
    subj_conf_data = etree.SubElement(subj_conf, '{urn:oasis:names:tc:SAML:2.0:assertion}SubjectConfirmationData')
    subj_conf_data.set('NotOnOrAfter', not_on_or_after_str)
    subj_conf_data.set('Recipient', acs_url)

    conditions_elem = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}Conditions')
    conditions_elem.set('NotBefore', now_str)
    conditions_elem.set('NotOnOrAfter', not_on_or_after_str)
    audience_restriction = etree.SubElement(conditions_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AudienceRestriction')
    audience_elem = etree.SubElement(audience_restriction, '{urn:oasis:names:tc:SAML:2.0:assertion}Audience')
    audience_elem.text = app.get('entity_id', '')

    authn_stmt = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnStatement')
    authn_stmt.set('AuthnInstant', now_str)
    authn_ctx = etree.SubElement(authn_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnContext')
    authn_ctx_ref = etree.SubElement(authn_ctx, '{urn:oasis:names:tc:SAML:2.0:assertion}AuthnContextClassRef')
    authn_ctx_ref.text = 'urn:oasis:names:tc:SAML:2.0:ac:classes:Password'

    attr_stmt = etree.SubElement(assertion_elem, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeStatement')
    email_attr = etree.SubElement(attr_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}Attribute')
    email_attr.set('Name', 'email')
    email_val = etree.SubElement(email_attr, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeValue')
    email_val.text = user.get('email', '')
    name_attr = etree.SubElement(attr_stmt, '{urn:oasis:names:tc:SAML:2.0:assertion}Attribute')
    name_attr.set('Name', 'name')
    name_val = etree.SubElement(name_attr, '{urn:oasis:names:tc:SAML:2.0:assertion}AttributeValue')
    name_val.text = user.get('name', '')

    # Sign the assertion (cert_pem and key_pem already set above, potentially from parent app)
    signed = False

    if key_pem and cert_pem:
        try:
            import signxml
            from signxml import XMLSigner
            signer = XMLSigner(
                method=signxml.methods.enveloped,
                signature_algorithm="rsa-sha256",
                digest_algorithm="sha256",
                c14n_algorithm="http://www.w3.org/2001/10/xml-exc-c14n#"
            )
            signed_assertion = signer.sign(assertion_elem, key=key_pem, cert=cert_pem, reference_uri=f'#{assertion_id}')
            for child in list(response_elem):
                if child.tag == '{urn:oasis:names:tc:SAML:2.0:assertion}Assertion':
                    response_elem.remove(child)
            response_elem.append(signed_assertion)
            signed = True
        except Exception as e:
            logging.error(f"SAML test signing failed: {e}")

    # Clean base64 text content in signature elements (remove PEM newlines)
    ds_ns = 'http://www.w3.org/2000/09/xmldsig#'
    for tag in ['X509Certificate', 'SignatureValue', 'DigestValue']:
        for elem in response_elem.iter(f'{{{ds_ns}}}{tag}'):
            if elem.text:
                elem.text = elem.text.replace('\n', '').replace('\r', '').replace(' ', '')

    xml_str = etree.tostring(response_elem, xml_declaration=False, encoding='unicode', pretty_print=False).strip()
    pretty_xml = etree.tostring(response_elem, pretty_print=True, xml_declaration=False, encoding='unicode')
    saml_b64 = base64.b64encode(xml_str.encode('utf-8')).decode('ascii')

    return {
        "status": "success",
        "signed": signed,
        "response_id": response_id,
        "assertion_id": assertion_id,
        "issuer": issuer,
        "destination": acs_url,
        "audience": app.get('entity_id', ''),
        "name_id": name_id,
        "name_id_format": name_id_format,
        "issue_instant": now_str,
        "not_on_or_after": not_on_or_after_str,
        "attributes": {"email": user.get('email', ''), "name": user.get('name', '')},
        "saml_response_b64": saml_b64,
        "saml_response_xml": pretty_xml,
        "acs_url": acs_url,
    }

# ===================== OIDC APP ROUTES =====================

@api_router.get("/apps/oidc")
async def list_oidc_apps(user: dict = Depends(get_current_user)):
    apps = await db.oidc_apps.find({"org_id": user['org_id']}, {"_id": 0, "client_secret": 0}).to_list(100)
    return apps

@api_router.post("/apps/oidc")
async def create_oidc_app(app: OIDCAppCreate, request: Request, user: dict = Depends(get_current_user)):
    if app.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    app_id = str(uuid.uuid4())
    client_id = f"oidc_{str(uuid.uuid4()).replace('-', '')[:16]}"
    client_secret = str(uuid.uuid4()).replace('-', '') + str(uuid.uuid4()).replace('-', '')
    base_url = get_public_base_url(request)
    
    app_doc = {
        "id": app_id,
        "name": app.name,
        "description": app.description,
        "org_id": app.org_id,
        "client_id": client_id,
        "client_secret": client_secret,
        "redirect_uris": app.redirect_uris,
        "logout_uris": app.logout_uris,
        "scopes": app.scopes,
        "grant_types": app.grant_types,
        "authorization_endpoint": f"{base_url}/api/oidc/{app_id}/authorize",
        "token_endpoint": f"{base_url}/api/oidc/{app_id}/token",
        "userinfo_endpoint": f"{base_url}/api/oidc/{app_id}/userinfo",
        "logo_url": app.logo_url,
        "home_url": app.home_url,
        "allowed_group_ids": app.allowed_group_ids,
        "allowed_role_ids": app.allowed_role_ids,
        "status": "active",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.oidc_apps.insert_one(app_doc)
    await log_audit(user['org_id'], "oidc_app_created", "app", user['id'], user['email'], app_id,
                   {"name": app.name}, request.client.host if request.client else None)
    
    return {k: v for k, v in app_doc.items() if k != '_id'}

@api_router.get("/apps/oidc/{app_id}")
async def get_oidc_app(app_id: str, include_secret: bool = False, user: dict = Depends(get_current_user)):
    projection = {"_id": 0} if include_secret else {"_id": 0, "client_secret": 0}
    app = await db.oidc_apps.find_one({"id": app_id, "org_id": user['org_id']}, projection)
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    return app

@api_router.put("/apps/oidc/{app_id}")
async def update_oidc_app(app_id: str, update: dict, request: Request, user: dict = Depends(get_current_user)):
    app = await db.oidc_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    update.pop('id', None)
    update.pop('org_id', None)
    update.pop('client_id', None)
    update.pop('client_secret', None)
    
    await db.oidc_apps.update_one({"id": app_id}, {"$set": update})
    await log_audit(user['org_id'], "oidc_app_updated", "app", user['id'], user['email'], app_id,
                   update, request.client.host if request.client else None)
    return await db.oidc_apps.find_one({"id": app_id}, {"_id": 0, "client_secret": 0})

@api_router.delete("/apps/oidc/{app_id}")
async def delete_oidc_app(app_id: str, request: Request, user: dict = Depends(get_current_user)):
    app = await db.oidc_apps.find_one({"id": app_id, "org_id": user['org_id']}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    await db.oidc_apps.delete_one({"id": app_id})
    await log_audit(user['org_id'], "oidc_app_deleted", "app", user['id'], user['email'], app_id,
                   {"name": app['name']}, request.client.host if request.client else None)
    return {"message": "App deleted"}

@api_router.get("/apps/oidc/{app_id}/.well-known/openid-configuration")
async def get_oidc_discovery(app_id: str, request: Request):
    app = await db.oidc_apps.find_one({"id": app_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=404, detail="App not found")
    
    base_url = get_public_base_url(request)
    return {
        "issuer": base_url,
        "authorization_endpoint": f"{base_url}/api/oidc/{app_id}/authorize",
        "token_endpoint": f"{base_url}/api/oidc/{app_id}/token",
        "userinfo_endpoint": f"{base_url}/api/oidc/userinfo",
        "jwks_uri": f"{base_url}/api/oidc/jwks",
        "scopes_supported": ["openid", "profile", "email"],
        "response_types_supported": ["code"],
        "grant_types_supported": ["authorization_code"],
        "subject_types_supported": ["public"],
        "id_token_signing_alg_values_supported": ["HS256"],
        "token_endpoint_auth_methods_supported": ["client_secret_post", "client_secret_basic"],
    }

# ===================== OIDC PROVIDER FLOW =====================

@api_router.get("/oidc/{app_id}/authorize")
async def oidc_authorize(
    app_id: str,
    request: Request,
    response_type: str = Query(...),
    client_id: str = Query(...),
    redirect_uri: str = Query(...),
    scope: str = Query("openid"),
    state: Optional[str] = Query(None),
    nonce: Optional[str] = Query(None),
):
    """OAuth2 Authorization endpoint - shows login page or redirects with auth code"""
    # Find the OIDC app by client_id
    app = await db.oidc_apps.find_one({"client_id": client_id, "id": app_id}, {"_id": 0})
    if not app:
        # Also try finding just by client_id (some apps don't include app_id in URL)
        app = await db.oidc_apps.find_one({"client_id": client_id}, {"_id": 0})
    if not app:
        raise HTTPException(status_code=400, detail="Invalid client_id")
    
    if app.get('status') != 'active':
        raise HTTPException(status_code=400, detail="Application is inactive")
    
    # Validate redirect_uri
    if redirect_uri not in app.get('redirect_uris', []):
        raise HTTPException(status_code=400, detail=f"Invalid redirect_uri. Allowed: {app.get('redirect_uris', [])}")
    
    if response_type != 'code':
        raise HTTPException(status_code=400, detail="Only response_type=code is supported")
    
    # Check if user has an active session (IAM token cookie or query param)
    token = request.query_params.get('token') or request.cookies.get('iam_token')
    if not token:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Bearer '):
            token = auth_header[7:]
    
    base_url = get_public_base_url(request)
    
    if token:
        try:
            payload = decode_token(token)
            user = await db.users.find_one({"id": payload['user_id']}, {"_id": 0})
            if user:
                # User is authenticated - generate authorization code
                auth_code = str(uuid.uuid4()).replace('-', '')
                
                # Store auth code in DB with expiry
                await db.oidc_auth_codes.insert_one({
                    "code": auth_code,
                    "client_id": client_id,
                    "app_id": app.get('id'),
                    "user_id": user['id'],
                    "email": user['email'],
                    "name": user.get('name', user.get('full_name', '')),
                    "org_id": user.get('org_id', ''),
                    "redirect_uri": redirect_uri,
                    "scope": scope,
                    "nonce": nonce,
                    "created_at": datetime.now(timezone.utc),
                    "expires_at": datetime.now(timezone.utc) + timedelta(minutes=10),
                    "used": False,
                })
                
                # Redirect back to the app with authorization code
                separator = '&' if '?' in redirect_uri else '?'
                redirect_url = f"{redirect_uri}{separator}code={auth_code}"
                if state:
                    redirect_url += f"&state={state}"
                
                return Response(
                    status_code=302,
                    headers={"Location": redirect_url}
                )
        except Exception:
            pass  # Token invalid, show login page
    
    # No valid session - show login page that will redirect back after auth
    # Build the authorize URL to come back to after login
    import urllib.parse
    params = {
        "response_type": response_type,
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "scope": scope,
    }
    if state:
        params["state"] = state
    if nonce:
        params["nonce"] = nonce
    
    authorize_url = f"{base_url}/api/oidc/{app_id}/authorize?{urllib.parse.urlencode(params)}"
    login_url = f"{base_url}/login?oidc_redirect={urllib.parse.quote(authorize_url)}"
    
    return Response(
        status_code=302,
        headers={"Location": login_url}
    )

@api_router.post("/oidc/{app_id}/token")
async def oidc_token(app_id: str, request: Request):
    """OAuth2 Token endpoint - exchanges authorization code for tokens"""
    # Parse form data or JSON body
    content_type = request.headers.get('content-type', '')
    if 'application/x-www-form-urlencoded' in content_type:
        form = await request.form()
        data = dict(form)
    elif 'application/json' in content_type:
        data = await request.json()
    else:
        # Try form first, fallback to JSON
        try:
            form = await request.form()
            data = dict(form)
        except:
            data = await request.json()
    
    grant_type = data.get('grant_type')
    code = data.get('code')
    redirect_uri = data.get('redirect_uri')
    client_id = data.get('client_id')
    client_secret = data.get('client_secret')
    
    # Support Basic Auth for client credentials
    if not client_id or not client_secret:
        auth_header = request.headers.get('Authorization', '')
        if auth_header.startswith('Basic '):
            import base64
            decoded = base64.b64decode(auth_header[6:]).decode('utf-8')
            client_id, client_secret = decoded.split(':', 1)
    
    if grant_type != 'authorization_code':
        return Response(
            content='{"error": "unsupported_grant_type"}',
            status_code=400,
            media_type="application/json"
        )
    
    if not code or not client_id or not client_secret:
        return Response(
            content='{"error": "invalid_request", "error_description": "Missing required parameters"}',
            status_code=400,
            media_type="application/json"
        )
    
    # Verify client credentials
    app = await db.oidc_apps.find_one({"client_id": client_id}, {"_id": 0})
    if not app or app.get('client_secret') != client_secret:
        return Response(
            content='{"error": "invalid_client"}',
            status_code=401,
            media_type="application/json"
        )
    
    # Verify authorization code
    auth_code = await db.oidc_auth_codes.find_one({"code": code, "client_id": client_id}, {"_id": 0})
    if not auth_code:
        return Response(
            content='{"error": "invalid_grant", "error_description": "Invalid authorization code"}',
            status_code=400,
            media_type="application/json"
        )
    
    if auth_code.get('used'):
        return Response(
            content='{"error": "invalid_grant", "error_description": "Authorization code already used"}',
            status_code=400,
            media_type="application/json"
        )
    
    expires_at = auth_code.get('expires_at')
    if isinstance(expires_at, str):
        expires_at = datetime.fromisoformat(expires_at)
    if expires_at:
        # Ensure timezone-aware comparison
        if expires_at.tzinfo is None:
            expires_at = expires_at.replace(tzinfo=timezone.utc)
        if expires_at < datetime.now(timezone.utc):
            return Response(
                content='{"error": "invalid_grant", "error_description": "Authorization code expired"}',
                status_code=400,
                media_type="application/json"
            )
    
    if redirect_uri and redirect_uri != auth_code.get('redirect_uri'):
        return Response(
            content='{"error": "invalid_grant", "error_description": "redirect_uri mismatch"}',
            status_code=400,
            media_type="application/json"
        )
    
    # Mark code as used
    await db.oidc_auth_codes.update_one({"code": code}, {"$set": {"used": True}})
    
    base_url = get_public_base_url(request)
    now = datetime.now(timezone.utc)
    
    # Generate access token
    access_token_payload = {
        "sub": auth_code['user_id'],
        "email": auth_code['email'],
        "name": auth_code.get('name', ''),
        "iss": base_url,
        "aud": client_id,
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(hours=1)).timestamp()),
        "scope": auth_code.get('scope', 'openid'),
    }
    access_token = jwt.encode(access_token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    # Generate ID token (for OpenID Connect)
    id_token_payload = {
        "sub": auth_code['user_id'],
        "email": auth_code['email'],
        "name": auth_code.get('name', ''),
        "iss": base_url,
        "aud": client_id,
        "iat": int(now.timestamp()),
        "exp": int((now + timedelta(hours=1)).timestamp()),
        "at_hash": "",  # Simplified
    }
    if auth_code.get('nonce'):
        id_token_payload['nonce'] = auth_code['nonce']
    id_token = jwt.encode(id_token_payload, JWT_SECRET, algorithm=JWT_ALGORITHM)
    
    # Store access token for userinfo lookups
    await db.oidc_access_tokens.insert_one({
        "access_token": access_token,
        "user_id": auth_code['user_id'],
        "email": auth_code['email'],
        "name": auth_code.get('name', ''),
        "org_id": auth_code.get('org_id', ''),
        "client_id": client_id,
        "scope": auth_code.get('scope', 'openid'),
        "created_at": now,
        "expires_at": now + timedelta(hours=1),
    })
    
    return {
        "access_token": access_token,
        "token_type": "Bearer",
        "expires_in": 3600,
        "id_token": id_token,
        "scope": auth_code.get('scope', 'openid'),
    }

@api_router.get("/oidc/userinfo")
@api_router.post("/oidc/userinfo")
async def oidc_userinfo(request: Request):
    """OIDC UserInfo endpoint - returns user profile from access token"""
    auth_header = request.headers.get('Authorization', '')
    if not auth_header.startswith('Bearer '):
        raise HTTPException(status_code=401, detail="Missing Bearer token")
    
    access_token = auth_header[7:]
    
    # Try to decode the JWT directly
    try:
        payload = jwt.decode(access_token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get('sub')
        user = await db.users.find_one({"id": user_id}, {"_id": 0})
        if user:
            response = {
                "sub": user['id'],
                "email": user.get('email', ''),
                "email_verified": True,
                "name": user.get('name', user.get('full_name', '')),
            }
            # Add profile fields if scope includes 'profile'
            if user.get('name') or user.get('full_name'):
                full_name = user.get('name', user.get('full_name', ''))
                parts = full_name.split(' ', 1)
                response['given_name'] = parts[0]
                response['family_name'] = parts[1] if len(parts) > 1 else ''
            
            # Google-compatible picture field
            if user.get('picture') or user.get('avatar_url'):
                response['picture'] = user.get('picture', user.get('avatar_url', ''))
            
            return response
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except Exception:
        pass
    
    # Fallback: look up in stored tokens
    token_doc = await db.oidc_access_tokens.find_one({"access_token": access_token}, {"_id": 0})
    if not token_doc:
        raise HTTPException(status_code=401, detail="Invalid access token")
    
    return {
        "sub": token_doc['user_id'],
        "email": token_doc.get('email', ''),
        "email_verified": True,
        "name": token_doc.get('name', ''),
    }

# ===================== ACCESS POLICY ROUTES =====================

@api_router.get("/policies")
async def list_policies(user: dict = Depends(get_current_user)):
    policies = await db.access_policies.find({"org_id": user['org_id']}, {"_id": 0}).to_list(100)
    return policies

@api_router.post("/policies")
async def create_policy(policy: AccessPolicyCreate, request: Request, user: dict = Depends(get_current_user)):
    if policy.org_id != user['org_id']:
        raise HTTPException(status_code=403, detail="Access denied")
    
    policy_id = str(uuid.uuid4())
    policy_doc = {
        "id": policy_id,
        "name": policy.name,
        "description": policy.description,
        "org_id": policy.org_id,
        "app_ids": policy.app_ids,
        "conditions": policy.conditions,
        "enabled": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    
    await db.access_policies.insert_one(policy_doc)
    await log_audit(user['org_id'], "policy_created", "policy", user['id'], user['email'], policy_id,
                   {"name": policy.name}, request.client.host if request.client else None)
    return {**policy_doc, "_id": None}

@api_router.put("/policies/{policy_id}")
async def update_policy(policy_id: str, update: dict, request: Request, user: dict = Depends(get_current_user)):
    policy = await db.access_policies.find_one({"id": policy_id, "org_id": user['org_id']}, {"_id": 0})
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    
    update.pop('id', None)
    update.pop('org_id', None)
    await db.access_policies.update_one({"id": policy_id}, {"$set": update})
    await log_audit(user['org_id'], "policy_updated", "policy", user['id'], user['email'], policy_id,
                   update, request.client.host if request.client else None)
    return await db.access_policies.find_one({"id": policy_id}, {"_id": 0})

@api_router.delete("/policies/{policy_id}")
async def delete_policy(policy_id: str, request: Request, user: dict = Depends(get_current_user)):
    policy = await db.access_policies.find_one({"id": policy_id, "org_id": user['org_id']}, {"_id": 0})
    if not policy:
        raise HTTPException(status_code=404, detail="Policy not found")
    
    await db.access_policies.delete_one({"id": policy_id})
    await log_audit(user['org_id'], "policy_deleted", "policy", user['id'], user['email'], policy_id,
                   {"name": policy['name']}, request.client.host if request.client else None)
    return {"message": "Policy deleted"}

# ===================== AUDIT LOG ROUTES =====================

@api_router.get("/audit-logs")
async def list_audit_logs(
    action: str = None,
    resource_type: str = None,
    user_id: str = None,
    start_date: str = None,
    end_date: str = None,
    limit: int = Query(default=100, le=500),
    user: dict = Depends(get_current_user)
):
    query = {"org_id": user['org_id']}
    
    if action:
        query['action'] = action
    if resource_type:
        query['resource_type'] = resource_type
    if user_id:
        query['user_id'] = user_id
    if start_date:
        query['timestamp'] = {"$gte": start_date}
    if end_date:
        if 'timestamp' in query:
            query['timestamp']['$lte'] = end_date
        else:
            query['timestamp'] = {"$lte": end_date}
    
    logs = await db.audit_logs.find(query, {"_id": 0}).sort("timestamp", -1).to_list(limit)
    return logs

@api_router.get("/audit-logs/summary")
async def get_audit_summary(user: dict = Depends(get_current_user)):
    org_id = user['org_id']
    
    # Get counts by action type
    pipeline = [
        {"$match": {"org_id": org_id}},
        {"$group": {"_id": "$action", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}}
    ]
    action_counts = await db.audit_logs.aggregate(pipeline).to_list(100)
    
    # Get recent activity
    recent = await db.audit_logs.find({"org_id": org_id}, {"_id": 0}).sort("timestamp", -1).to_list(10)
    
    # Get unique users
    unique_users = await db.audit_logs.distinct("user_id", {"org_id": org_id})
    
    return {
        "action_counts": {item['_id']: item['count'] for item in action_counts},
        "recent_activity": recent,
        "unique_users": len(unique_users),
        "total_logs": await db.audit_logs.count_documents({"org_id": org_id})
    }

# ===================== APP LAUNCHER / CATALOG ROUTES =====================

@api_router.get("/launcher/apps")
async def get_user_apps(request: Request, user: dict = Depends(get_current_user)):
    """Get all apps the user has access to"""
    org_id = user['org_id']
    
    # Get all SAML apps
    saml_apps = await db.saml_apps.find({"org_id": org_id, "status": "active"}, {"_id": 0, "private_key": 0, "certificate": 0}).to_list(100)
    # Get all OIDC apps  
    oidc_apps = await db.oidc_apps.find({"org_id": org_id, "status": "active"}, {"_id": 0, "client_secret": 0}).to_list(100)
    
    accessible_apps = []
    
    for app in saml_apps:
        has_access = await check_user_app_access(user, app)
        if has_access:
            allowed, reason = await check_access_policies(user, app, request)
            accessible_apps.append({
                "id": app['id'],
                "name": app['name'],
                "description": app.get('description'),
                "logo_url": app.get('logo_url'),
                "home_url": app.get('home_url'),
                "type": "saml",
                "launch_url": f"/api/saml/{app['id']}/sso",
                "policy_blocked": not allowed,
                "policy_reason": reason
            })
    
    for app in oidc_apps:
        has_access = await check_user_app_access(user, app)
        if has_access:
            allowed, reason = await check_access_policies(user, app, request)
            accessible_apps.append({
                "id": app['id'],
                "name": app['name'],
                "description": app.get('description'),
                "logo_url": app.get('logo_url'),
                "home_url": app.get('home_url'),
                "type": "oidc",
                "launch_url": f"/api/oidc/{app['id']}/authorize",
                "policy_blocked": not allowed,
                "policy_reason": reason
            })
    
    return accessible_apps

@api_router.get("/catalog/apps")
async def get_app_catalog(user: dict = Depends(get_current_user)):
    """Get all apps in the catalog (for requesting access)"""
    org_id = user['org_id']
    
    saml_apps = await db.saml_apps.find({"org_id": org_id, "status": "active"}, 
                                         {"_id": 0, "id": 1, "name": 1, "description": 1, "logo_url": 1, 
                                          "allowed_group_ids": 1, "allowed_role_ids": 1, "approved_user_ids": 1}).to_list(100)
    oidc_apps = await db.oidc_apps.find({"org_id": org_id, "status": "active"},
                                         {"_id": 0, "id": 1, "name": 1, "description": 1, "logo_url": 1,
                                          "allowed_group_ids": 1, "allowed_role_ids": 1, "approved_user_ids": 1}).to_list(100)
    
    catalog = []
    
    for app in saml_apps:
        has_access = await check_user_app_access(user, app)
        # requires_approval is always true - explicit assignment needed
        catalog.append({
            "id": app['id'],
            "name": app['name'],
            "description": app.get('description'),
            "logo_url": app.get('logo_url'),
            "type": "saml",
            "has_access": has_access,
            "requires_approval": True
        })
    
    for app in oidc_apps:
        has_access = await check_user_app_access(user, app)
        catalog.append({
            "id": app['id'],
            "name": app['name'],
            "description": app.get('description'),
            "logo_url": app.get('logo_url'),
            "type": "oidc",
            "has_access": has_access,
            "requires_approval": True
        })
    
    return catalog

# ===================== DASHBOARD STATS =====================

@api_router.get("/dashboard/stats")
async def get_dashboard_stats(user: dict = Depends(get_current_user)):
    org_id = user['org_id']
    
    return {
        "total_users": await db.users.count_documents({"org_id": org_id}),
        "active_users": await db.users.count_documents({"org_id": org_id, "status": "active"}),
        "total_groups": await db.groups.count_documents({"org_id": org_id}),
        "total_roles": await db.roles.count_documents({"org_id": org_id}),
        "saml_apps": await db.saml_apps.count_documents({"org_id": org_id}),
        "oidc_apps": await db.oidc_apps.count_documents({"org_id": org_id}),
        "access_policies": await db.access_policies.count_documents({"org_id": org_id}),
        "pending_requests": await db.access_requests.count_documents({"org_id": org_id, "status": "pending"}),
        "recent_logins": await db.audit_logs.count_documents({
            "org_id": org_id, 
            "action": "user_login",
            "timestamp": {"$gte": (datetime.now(timezone.utc) - timedelta(days=7)).isoformat()}
        })
    }

# ===================== BASE ROUTES =====================

@api_router.get("/")
async def root():
    return {"message": "Kissflow IAM - Identity & Access Management API"}

# ===================== SCIM TOKEN MANAGEMENT =====================

@api_router.post("/scim/tokens")
async def create_scim_token(body: dict, user: dict = Depends(get_current_user)):
    """Admin generates a SCIM bearer token for external clients"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can manage SCIM tokens")

    label = body.get("label", "SCIM Token")
    token_value = f"scim_{uuid.uuid4().hex}{uuid.uuid4().hex}"

    token_doc = {
        "id": str(uuid.uuid4()),
        "token": token_value,
        "label": label,
        "org_id": user["org_id"],
        "created_by": user["id"],
        "active": True,
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.scim_tokens.insert_one(token_doc)

    base_url = get_public_base_url()
    return {
        "id": token_doc["id"],
        "token": token_value,
        "label": label,
        "scim_base_url": f"{base_url}/api/scim/v2",
        "created_at": token_doc["created_at"],
    }


@api_router.get("/scim/tokens")
async def list_scim_tokens(user: dict = Depends(get_current_user)):
    """List active SCIM tokens for the org"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can view SCIM tokens")

    tokens = await db.scim_tokens.find(
        {"org_id": user["org_id"], "active": True},
        {"_id": 0, "token": 0}  # Hide actual token value in listings
    ).to_list(50)

    base_url = get_public_base_url()
    for t in tokens:
        t["scim_base_url"] = f"{base_url}/api/scim/v2"

    return tokens


@api_router.delete("/scim/tokens/{token_id}")
async def revoke_scim_token(token_id: str, user: dict = Depends(get_current_user)):
    """Revoke a SCIM token"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can manage SCIM tokens")

    result = await db.scim_tokens.update_one(
        {"id": token_id, "org_id": user["org_id"]},
        {"$set": {"active": False, "revoked_at": datetime.now(timezone.utc).isoformat()}}
    )
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Token not found")

    return {"message": "Token revoked"}


# ===================== ACCESS REQUESTS =====================

@api_router.post("/access-requests")
async def create_access_request(body: dict, user: dict = Depends(get_current_user)):
    """User requests access to an app"""
    app_id = body.get("app_id")
    app_type = body.get("app_type", "saml")  # saml or oidc
    reason = body.get("reason", "")

    if not app_id:
        raise HTTPException(status_code=400, detail="app_id is required")

    # Check if already requested
    existing = await db.access_requests.find_one({
        "user_id": user["id"], "app_id": app_id, "status": "pending"
    })
    if existing:
        raise HTTPException(status_code=400, detail="You already have a pending request for this app")

    # Get app details
    collection = "saml_apps" if app_type == "saml" else "oidc_apps"
    app_doc = await db[collection].find_one({"id": app_id}, {"_id": 0, "name": 1, "id": 1})
    if not app_doc:
        raise HTTPException(status_code=404, detail="Application not found")

    request_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user.get("name", user.get("full_name", "")),
        "user_email": user["email"],
        "app_id": app_id,
        "app_type": app_type,
        "app_name": app_doc.get("name", ""),
        "reason": reason,
        "org_id": user["org_id"],
        "status": "pending",
        "created_at": datetime.now(timezone.utc).isoformat(),
    }
    await db.access_requests.insert_one(request_doc)

    # Send email to all admins in the org
    admins = await db.users.find(
        {"org_id": user["org_id"], "role": "org_admin", "status": {"$ne": "disabled"}},
        {"_id": 0, "email": 1}
    ).to_list(100)
    admin_emails = [a["email"] for a in admins if a.get("email")]

    if admin_emails:
        base_url = get_public_base_url()
        approve_url = f"{base_url}/access-requests"
        html = build_access_request_email(
            request_doc["user_name"], request_doc["user_email"],
            request_doc["app_name"], reason, approve_url
        )
        await send_email(admin_emails, f"Access Request: {request_doc['app_name']} - {request_doc['user_name']}", html)

    request_doc.pop("_id", None)
    return request_doc


@api_router.get("/access-requests")
async def list_access_requests(status: str = None, user: dict = Depends(get_current_user)):
    """List access requests - admins see all for org, users see their own"""
    query = {"org_id": user["org_id"]}
    if user.get("role") != "org_admin":
        query["user_id"] = user["id"]
    if status:
        query["status"] = status

    requests = await db.access_requests.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    return requests


@api_router.put("/access-requests/{request_id}")
async def update_access_request(request_id: str, body: dict, user: dict = Depends(get_current_user)):
    """Admin approves or rejects an access request"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can approve/reject requests")

    new_status = body.get("status")
    if new_status not in ["approved", "rejected"]:
        raise HTTPException(status_code=400, detail="Status must be 'approved' or 'rejected'")

    req = await db.access_requests.find_one({"id": request_id, "org_id": user["org_id"]}, {"_id": 0})
    if not req:
        raise HTTPException(status_code=404, detail="Request not found")
    if req["status"] != "pending":
        raise HTTPException(status_code=400, detail="Request already processed")

    await db.access_requests.update_one(
        {"id": request_id},
        {"$set": {
            "status": new_status,
            "reviewed_by": user["id"],
            "reviewed_by_name": user.get("name", ""),
            "reviewed_at": datetime.now(timezone.utc).isoformat(),
        }}
    )

    # If approved, add user to the app's approved_user_ids
    if new_status == "approved":
        collection = "saml_apps" if req.get("app_type") == "saml" else "oidc_apps"
        await db[collection].update_one(
            {"id": req["app_id"]},
            {"$addToSet": {"approved_user_ids": req["user_id"]}}
        )

    # Send status email to the requester
    requester = await db.users.find_one({"id": req["user_id"]}, {"_id": 0, "email": 1, "name": 1})
    if requester and requester.get("email"):
        html = build_request_status_email(
            requester.get("name", "User"), req["app_name"],
            new_status, user.get("name", "Admin")
        )
        status_word = "Approved" if new_status == "approved" else "Rejected"
        await send_email([requester["email"]], f"Access {status_word}: {req['app_name']}", html)

    return {"message": f"Request {new_status}", "id": request_id}


# ===================== HR SYNC (ADRENALIN) =====================

@api_router.post("/hr-sync/trigger")
async def trigger_hr_sync(user: dict = Depends(get_current_user)):
    """Admin manually triggers HR sync"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can trigger HR sync")

    result = await sync_employees(db, user["org_id"])

    # Log the sync
    log_doc = {
        "org_id": user["org_id"],
        "triggered_by": user["id"],
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "result": result,
    }
    await db.hr_sync_logs.insert_one(log_doc)

    # Enable adrenalin sync for this org
    await db.organizations.update_one(
        {"id": user["org_id"]},
        {"$set": {"adrenalin_sync_enabled": True}}
    )

    return result


@api_router.get("/hr-sync/logs")
async def get_hr_sync_logs(user: dict = Depends(get_current_user)):
    """Get HR sync history"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can view sync logs")

    logs = await db.hr_sync_logs.find(
        {"org_id": user["org_id"]}, {"_id": 0}
    ).sort("timestamp", -1).to_list(50)
    return logs


# ---- Kissflow SCIM Push Endpoints ----

@api_router.get("/kissflow-scim/config")
async def get_kf_scim_config(user: dict = Depends(get_current_user)):
    """Get Kissflow SCIM configuration status"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Admin only")
    config = await get_kissflow_scim_config(db, user["org_id"])
    if not config:
        return {"configured": False}
    return {
        "configured": True,
        "base_url": config.get("base_url", ""),
        "token_masked": config.get("token", "")[:8] + "..." if config.get("token") else "",
        "source": config.get("source", "db"),
    }


@api_router.post("/kissflow-scim/config")
async def save_kf_scim_config(request: Request, user: dict = Depends(get_current_user)):
    """Save Kissflow SCIM configuration"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Admin only")
    body = await request.json()
    base_url = body.get("base_url", "").strip()
    token = body.get("token", "").strip()
    if not base_url or not token:
        raise HTTPException(status_code=400, detail="base_url and token are required")
    await save_kissflow_scim_config(db, user["org_id"], base_url, token)
    return {"success": True}


@api_router.post("/kissflow-scim/sync")
async def trigger_kissflow_sync(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    """Admin manually triggers Kissflow SCIM push for all users (runs in background)"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can trigger Kissflow sync")

    org_id = user["org_id"]
    user_id = user["id"]

    # Create a pending log entry immediately
    log_doc = {
        "org_id": org_id,
        "triggered_by": user_id,
        "trigger_type": "manual_full",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "running",
        "result": {"message": "Sync started in background..."},
    }
    await db.kissflow_sync_logs.insert_one(log_doc)
    log_id = log_doc.get("_id")

    async def _run_sync():
        try:
            result = await sync_to_kissflow(db, org_id)
            await db.kissflow_sync_logs.update_one(
                {"_id": log_id},
                {"$set": {"result": result, "status": "completed"}}
            )
        except Exception as e:
            await db.kissflow_sync_logs.update_one(
                {"_id": log_id},
                {"$set": {"result": {"error": str(e)}, "status": "failed"}}
            )

    # Run in background so the HTTP response returns immediately
    background_tasks.add_task(_run_sync)

    return {"message": "Kissflow SCIM sync started in background. Check Sync History for progress.", "status": "running"}


@api_router.post("/kissflow-scim/push-user")
async def push_user_to_kf(request: Request, user: dict = Depends(get_current_user)):
    """Push a single user to Kissflow (used during admin edits)"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Admin only")
    body = await request.json()
    email = body.get("email", "").strip()
    if not email:
        raise HTTPException(status_code=400, detail="email is required")

    result = await push_single_user_to_kissflow(db, user["org_id"], email)

    log_doc = {
        "org_id": user["org_id"],
        "triggered_by": user["id"],
        "trigger_type": "manual_single",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "email": email,
        "result": result,
    }
    await db.kissflow_sync_logs.insert_one(log_doc)

    return result


@api_router.get("/kissflow-scim/logs")
async def get_kissflow_sync_logs(user: dict = Depends(get_current_user)):
    """Get Kissflow SCIM sync history"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Admin only")
    logs = await db.kissflow_sync_logs.find(
        {"org_id": user["org_id"]}, {"_id": 0}
    ).sort("timestamp", -1).to_list(50)
    return logs


@api_router.post("/kissflow-scim/resolve-managers")
async def trigger_manager_resolution(background_tasks: BackgroundTasks, user: dict = Depends(get_current_user)):
    """Second pass: resolve Manager/L2_Manager lookup fields with Kissflow User IDs"""
    if user.get("role") != "org_admin":
        raise HTTPException(status_code=403, detail="Only admins can trigger manager resolution")

    org_id = user["org_id"]

    log_doc = {
        "org_id": org_id,
        "triggered_by": user["id"],
        "trigger_type": "resolve_managers",
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "status": "running",
        "result": {"message": "Resolving manager lookups..."},
    }
    await db.kissflow_sync_logs.insert_one(log_doc)
    log_id = log_doc.get("_id")

    async def _run_resolve():
        try:
            result = await resolve_managers_in_kissflow(db, org_id)
            await db.kissflow_sync_logs.update_one(
                {"_id": log_id},
                {"$set": {"result": result, "status": "completed"}}
            )
        except Exception as e:
            await db.kissflow_sync_logs.update_one(
                {"_id": log_id},
                {"$set": {"result": {"error": str(e)}, "status": "failed"}}
            )

    background_tasks.add_task(_run_resolve)
    return {"message": "Manager resolution started in background.", "status": "running"}




@api_router.get("/health")
async def health():
    return {"status": "healthy", "timestamp": datetime.now(timezone.utc).isoformat()}

# Include router
app.include_router(api_router)

# SCIM v2 router (separate module)
scim_router_module.db = db
app.include_router(scim_router_module.router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
